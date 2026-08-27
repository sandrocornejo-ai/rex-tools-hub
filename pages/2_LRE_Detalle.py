import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import io
from datetime import datetime
import calendar
from modulo_dt import render_modulo_dt

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
DATA_DIR = "data"  # Carpeta con archivos de referencia

# ─────────────────────────────────────────────
# ESTILOS REX+
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Rex+ | Liquidaciones en detalle desde LRE",
    page_icon="💼",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Header */
.rex-header {
    background-color: #1a2744;
    padding: 14px 28px;
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 28px;
}
.rex-logo {
    background: white;
    color: #1a2744;
    font-weight: 800;
    font-size: 15px;
    padding: 5px 10px;
    border-radius: 6px;
    letter-spacing: 0.5px;
}
.rex-logo span { color: #00b4d8; }
.rex-title { color: white; font-size: 18px; font-weight: 600; margin-left: 16px; }
.rex-badge {
    background: #00b4d8;
    color: white;
    font-size: 11px;
    font-weight: 700;
    padding: 4px 12px;
    border-radius: 20px;
    letter-spacing: 1px;
}

/* Cards */
.step-card {
    background: white;
    border: 1px solid #e8edf5;
    border-radius: 10px;
    padding: 18px 20px;
    margin-bottom: 12px;
}
.step-label {
    color: #00b4d8;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.step-title { font-size: 15px; font-weight: 600; color: #1a2744; margin-bottom: 4px; }
.step-desc { font-size: 13px; color: #6b7a9a; }

/* Section title */
.section-title {
    font-size: 20px;
    font-weight: 700;
    color: #1a2744;
    margin-bottom: 4px;
}
.section-sub { font-size: 13px; color: #6b7a9a; margin-bottom: 20px; }

/* Alerts */
.alert-error {
    background: #fff0f0;
    border-left: 4px solid #e53e3e;
    border-radius: 6px;
    padding: 12px 16px;
    margin: 8px 0;
    font-size: 13px;
    color: #c53030;
}
.alert-success {
    background: #f0fff4;
    border-left: 4px solid #38a169;
    border-radius: 6px;
    padding: 12px 16px;
    margin: 8px 0;
    font-size: 13px;
    color: #276749;
}
.alert-warning {
    background: #fffbf0;
    border-left: 4px solid #d69e2e;
    border-radius: 6px;
    padding: 12px 16px;
    margin: 8px 0;
    font-size: 13px;
    color: #744210;
}

/* Divider */
.rex-divider {
    border: none;
    border-top: 1px solid #e8edf5;
    margin: 24px 0;
}

/* Log table */
.log-header {
    background: #1a2744;
    color: white;
    font-size: 13px;
    font-weight: 600;
    padding: 10px 14px;
    border-radius: 8px 8px 0 0;
}

/* Buttons override */
.stButton > button {
    background-color: #1a2744 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    padding: 8px 20px !important;
    font-size: 14px !important;
}
.stButton > button:hover {
    background-color: #00b4d8 !important;
    color: white !important;
}

/* Sección parámetros */
.param-title {
    font-size: 0.8rem;
    font-weight: 700;
    color: #1a2744;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-bottom: 1rem;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid #00b4d8;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES DE COLUMNAS
# ─────────────────────────────────────────────
COLS_HABERES_AFECTOS = [
    "Sueldo(2101)", "Sobresueldo(2102)", "Comisiones(2103)", "Semana corrida(2104)",
    "Participación(2105)", "Gratificación(2106)", "Recargo 30% día domingo(2107)",
    "Remun. variable pagada en vacaciones(2108)", "Remun. variable pagada en clausura(2109)",
    "Aguinaldo(2110)", "Bonos u otras remun. fijas mensuales(2111)", "Tratos(2112)",
    "Bonos u otras remun. variables mensuales o superiores a un mes(2113)",
    "Ejercicio opción no pactada en contrato(2114)",
    "Beneficios en especie constitutivos de remun(2115)",
    "Remuneraciones bimestrales(2116)", "Remuneraciones trimestrales(2117)",
    "Remuneraciones cuatrimestral(2118)", "Remuneraciones semestrales(2119)",
    "Remuneraciones anuales(2120)", "Participación anual(2121)",
    "Gratificación anual(2122)", "Otras remuneraciones superiores a un mes(2123)",
    "Pago por horas de trabajo sindical(2124)", "Sueldo empresarial (2161)",
    "Subsidio por incapacidad laboral por licencia médica(2201)",
    "Beca de estudio(2202)", "Gratificaciones de zona(2203)"
]

COLS_HABERES_EXENTOS = [
    "Otros ingresos no constitutivos de renta(2204)", "Colación(2301)",
    "Movilización(2302)", "Viáticos(2303)", "Asignación de pérdida de caja(2304)",
    "Asignación de desgaste herramienta(2305)", "Asignación familiar legal(2311)",
    "Gastos por causa del trabajo(2306)", "Gastos por cambio de residencia(2307)",
    "Sala cuna(2308)", "Asignación trabajo a distancia o teletrabajo(2309)",
    "Depósito convenido hasta UF 900(2347)", "Alojamiento por razones de trabajo(2310)",
    "Asignación de traslación(2312)", "Indemnización por feriado legal(2313)",
    "Indemnización años de servicio(2314)", "Indemnización sustitutiva del aviso previo(2315)",
    "Indemnización fuero maternal(2316)", "Pago indemnización a todo evento(2331)",
    "Indemnizaciones voluntarias tributables(2417)",
    "Indemnizaciones contractuales tributables(2418)"
]

COLS_DESCUENTOS_LEGALES = [
    "Cotización obligatoria previsional (AFP o IPS)(3141)",
    "Cotización obligatoria salud 7%(3143)", "Cotización voluntaria para salud(3144)",
    "Cotización AFC - trabajador(3151)",
    "Cotizaciones técnico extranjero para seguridad social fuera de Chile(3146)",
    "Descuento depósito convenido hasta UF 900 anual(3147)",
    "Cotización APVi Mod A(3155)", "Cotización APVi Mod B hasta UF50(3156)",
    "Cotización APVc Mod A(3157)", "Cotización APVc Mod B hasta UF50(3158)",
    "Impuesto retenido por remuneraciones(3161)",
    "Impuesto retenido por indemnizaciones(3162)",
    "Mayor retención de impuestos solicitada por el trabajador(3163)",
    "Impuesto retenido por reliquidación remun. devengadas otros períodos(3164)",
    "Diferencia impuesto reliquidación remun. devengadas en este período(3165)",
    "Retención préstamo clase media 2020 (Ley 21.252) (3166)"
]

COLS_OTROS_DESCUENTOS = [
    "Cuota sindical 1(3171)", "Cuota sindical 2(3172)", "Cuota sindical 3(3173)",
    "Cuota sindical 4(3174)", "Cuota sindical 5(3175)", "Cuota sindical 6(3176)",
    "Cuota sindical 7(3177)", "Cuota sindical 8(3178)", "Cuota sindical 9(3179)",
    "Cuota sindical 10(3180)", "Crédito social CCAF(3110)",
    "Cuota vivienda o educación(3181)", "Crédito cooperativas de ahorro(3182)",
    "Otros descuentos autorizados y solicitados por el trabajador(3183)",
    "Cotización adicional trabajo pesado - trabajador(3154)",
    "Donaciones culturales y de reconstrucción(3184)", "Otros descuentos(3185)",
    "Pensiones de alimentos(3186)", "Descuento mujer casada(3187)",
    "Descuentos por anticipos y préstamos(3188)"
]

COLS_APORTES_EMPLEADOR = [
    "AFC - Aporte empleador(4151)",
    "Aporte empleador seguro accidentes del trabajo y Ley SANNA(4152)",
    "Aporte empleador indemnización a todo evento(4131)",
    "Aporte adicional trabajo pesado - empleador(4154)",
    "Aporte empleador seguro invalidez y sobrevivencia(4155)",
    "APVC - Aporte Empleador(4157)"
]

# Mapeo de columnas formato Rex+ → formato LRE esperado por el programa
COLS_REXPLUS_TO_LRE = {
    "Id empleado":                          "Rut trabajador (1101)",
    "Nro días trabajados":                  "Nro días trabajados en el mes(1115)",
    "Nro días de licencia médica":          "Nro días de licencia médica en el mes(1116)",
    "Recargo 30% día domingo (Art. 38) (2107)":                                             "Recargo 30% día domingo(2107)",
    "Remuneración variable pagada en vacaciones (Art 71) (cód 2108)":                       "Remun. variable pagada en vacaciones(2108)",
    "Tratos (mensual) (cód 2112)":          "Tratos(2112)",
    "Bonos u otras remuneraciones variables mensuales o superiores a un mes (cód 2113)":    "Bonos u otras remun. variables mensuales o superiores a un mes(2113)",
    "Beneficios en especie constitutivos de remuneración (cód 2115)":                       "Beneficios en especie constitutivos de remun(2115)",
    "Otras remuneraciones superiores a un mes (cód 2123)":                                  "Otras remuneraciones superiores a un mes(2123)",
    "Pago por horas de trabajo sindical (cód 2124)":                                        "Pago por horas de trabajo sindical(2124)",
    "Beca de estudio (Art. 17 N°18 LIR) (cód 2202)":                                       "Beca de estudio(2202)",
    "Otros ingresos no constitutivos de renta (Art 17 N°29 LIR) (cód 2204)":               "Otros ingresos no constitutivos de renta(2204)",
    "Viáticos totales mensual (Art 41) (cód 2303)":                                         "Viáticos(2303)",
    "Gastos por causa del trabajo (Art 41 CdT) y gastos de representación (Art. 42 Nº1 LIR) (cód 2306)": "Gastos por causa del trabajo(2306)",
    "Sala cuna (Art 203) (cód 2308)":       "Sala cuna(2308)",
    "Alojamiento por razones de trabajo (2310)":                                            "Alojamiento por razones de trabajo(2310)",
    "Indemnización fuero maternal (Art 163 bis) (cód 2316)":                                "Indemnización fuero maternal(2316)",
    "Indemnización a todo evento (Art.164) (cód 2331)":                                     "Pago indemnización a todo evento(2331)",
    "Indemnizaciones voluntarias tributables (cód 2417)":                                   "Indemnizaciones voluntarias tributables(2417)",
    "Indemnizaciones contractuales tributables (cód 2418)":                                 "Indemnizaciones contractuales tributables(2418)",
    "Cotización adicional trabajo pesado- trabajador (cód 3154)":                           "Cotización adicional trabajo pesado - trabajador(3154)",
    "Impuesto retenido por indemnizaciones (cód 3162)":                                     "Impuesto retenido por indemnizaciones(3162)",
    "Mayor retención de impuesto solicitada por el trabajador (cód 3163)":                  "Mayor retención de impuestos solicitada por el trabajador(3163)",
    "Impuesto retenido por reliquidación de remuneraciones devengadas en otros períodos mensuales (cód 3164)": "Impuesto retenido por reliquidación remun. devengadas otros períodos(3164)",
    "Cuota vivienda o educación Art. 58 (cód 3181)":                                        "Cuota vivienda o educación(3181)",
    "Crédito cooperativas de ahorro (Art 54 Ley Coop.) (cód 3182)":                         "Crédito cooperativas de ahorro(3182)",
    "Otros descuentos autorizados y solicitados por el trabajador (cód 3183)":              "Otros descuentos autorizados y solicitados por el trabajador(3183)",
    "Aporte adicional trabajo pesado- empleador (cód 4154)":                                "Aporte adicional trabajo pesado - empleador(4154)",
    "Mutual":                               "Org. administrador ley 16.744(1152)",
}

GRUPOS_AFP = {
    "afp", "reliquidaAfp", "afpAhor", "cesEmpleado", "reliquidaCesEmpl",
    "trabajoPesaEmpl", "voluntarioCoti", "voluntarioAhor", "reliquidaTrabEmpl",
    "trabajoPesa", "reliquidaTrabPesa", "sis", "reliquidaSis", "cesAporteCi",
    "reliquidaCesCi", "cesAporteSol", "reliquidaCesSol", "aporteAFPemp"
}
GRUPOS_ISAPRE = {"isapre", "reliquidaIsapre"}
GRUPOS_MUTUAL = {"mutual", "reliquidaMutual"}
GRUPOS_CCAF = {
    "ccafReliquida", "cajaCred", "cajaDent", "cajaLeas", "cajaVida",
    "cajaOtro", "cajaAhor", "cajaSegu", "cajaComp", "reliquidaCcaf"
}
GRUPOS_AFP_MUTUAL_AFECTO = {
    "afp", "isapre", "reliquidaIsapre", "reliquidaAfp", "afpAhor",
    "trabajoPesaEmpl", "voluntarioCoti", "voluntarioAhor", "reliquidaTrabEmpl",
    "trabajoPesa", "reliquidaTrabPesa", "sis", "reliquidaSis", "aporteAFPemp",
    "mutual", "reliquidaMutual"
}
GRUPOS_CES_AFECTO = {
    "cesEmpleado", "reliquidaCesEmpl", "cesAporteCi",
    "reliquidaCesCi", "cesAporteSol", "reliquidaCesSol"
}

# Conceptos que NO deben tener código LRE asignado en equiv_conceptos
CONCEPTOS_SIN_LRE = {
    "cajaComp", "reliquidaCcaf", "aporteAFPemp", "reliquidaAporteAFP",
    "aporteFAPPCEV", "reliquidaAporteCEV", "aporteFAPPBAC", "reliquidaAporteBAC",
    "aportesegurocovid",
}

# ─────────────────────────────────────────────
# CONSTANTES PARÁMETROS MENSUALES
# ─────────────────────────────────────────────
ARCHIVO_PARAMS = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
HOJA_PARAMS = "Hoja1"
LABELS_PARAMS = {
    "mes_Proc":           "Mes de proceso (aaaa-mm)",
    "uf_Mes":             "UF del mes ($)",
    "topeImp_Uf_afp":     "Tope imponible AFP (UF)",
    "topeImp_pesos_afp":  "Tope imponible AFP ($)",
    "topeCes_Uf":         "Tope cesantía (UF)",
    "topeCes_pesos":      "Tope cesantía ($)",
    "sis":                "SIS (%)",
    "factor_sis":         "Factor SIS (decimal)",
    "topeSalud_Uf":       "Tope salud (UF)",
    "topeSalud_pesos":    "Tope salud ($)",
    "imm":                "IMM ($)",
    "topeGratif":         "Tope gratificación ($)",
    "monto_Utm":          "Monto UTM ($)",
    "ult_Diames":         "Último día del mes",
    "aporte_Ccaf":        "Aporte CCAF (%)",
    "aporte_Fonasa":      "Aporte FONASA (%)",
    "Formato Fecha":      "Fecha formato (dd/mm/aaaa)",
    "Aporte AFP":         "Aporte AFP (%)",
    "Seg Social Exp vida":"Seg. social / Exp. vida (%)",
}

# ─────────────────────────────────────────────
# VALIDACIÓN DE CONCEPTOS PROHIBIDOS EN EQUIV
# ─────────────────────────────────────────────
def verificar_conceptos_prohibidos(equiv_df):
    """
    Revisa si alguno de los conceptos que NO deben tener código LRE
    aparece en equiv_conceptos con un cod_lre asignado.
    Retorna lista de dicts {Concepto, Código LRE asignado}.
    """
    if equiv_df is None or equiv_df.empty:
        return []
    if "concepto_detalle" not in equiv_df.columns or "cod_lre" not in equiv_df.columns:
        return []
    hallazgos = []
    for _, row in equiv_df.iterrows():
        concepto = str(row["concepto_detalle"]).strip()
        cod_lre  = str(row["cod_lre"]).strip()
        if concepto in CONCEPTOS_SIN_LRE and cod_lre and cod_lre.lower() != "nan":
            hallazgos.append({"Concepto": concepto, "Código LRE asignado": cod_lre})
    return hallazgos

def mostrar_alerta_conceptos_prohibidos(hallazgos):
    """Muestra en pantalla la advertencia cuando hay conceptos con código LRE no permitido."""
    if not hallazgos:
        return
    filas_html = "".join(
        f"<tr>"
        f"<td style='padding:4px 14px;font-weight:600;color:#744210'>{h['Concepto']}</td>"
        f"<td style='padding:4px 14px;color:#744210'>{h['Código LRE asignado']}</td>"
        f"</tr>"
        for h in hallazgos
    )
    st.markdown(f"""
    <div class="alert-warning">
        ⚠️ <b>¡Atención! {len(hallazgos)} concepto(s) tienen un código LRE asignado pero NO deberían tenerlo.</b><br>
        Por favor elimina esa relación en el archivo <b>equiv_conceptos.xlsx</b> antes de continuar:
        <table style='margin-top:10px;border-collapse:collapse;width:auto'>
            <thead>
                <tr>
                    <th style='padding:4px 14px;text-align:left;border-bottom:2px solid #d69e2e;color:#744210'>Concepto</th>
                    <th style='padding:4px 14px;text-align:left;border-bottom:2px solid #d69e2e;color:#744210'>Código LRE asignado</th>
                </tr>
            </thead>
            <tbody>{filas_html}</tbody>
        </table>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# FUNCIONES DE CARGA DE REFERENCIAS
# ─────────────────────────────────────────────
@st.cache_data
def cargar_referencias():
    refs = {}
    archivos = {
        "equiv_conceptos": "equiv_conceptos.xlsx",
        "inst_mutuales": "inst_mutuales.xlsx",
        "inst_cajas": "inst_cajas.xlsx",
        "inst_afp": "inst_afp.xlsx",
        "inst_salud": "inst_salud.xlsx",
        "cot_afp_hist": "cot_afp_hist.xlsx",
        "parametros": "parametrosMesuales.xlsx",
    }
    errores = []
    for key, fname in archivos.items():
        path = os.path.join(DATA_DIR, fname)
        if os.path.exists(path):
            refs[key] = pd.read_excel(path)
        else:
            errores.append(fname)
    return refs, errores

@st.cache_data(ttl=0)
def cargar_params():
    try:
        df = pd.read_excel(ARCHIVO_PARAMS, sheet_name=HOJA_PARAMS, dtype={"mes_Proc": str})
    except Exception:
        # Fallback: leer la primera hoja si el nombre no coincide
        df = pd.read_excel(ARCHIVO_PARAMS, sheet_name=0, dtype={"mes_Proc": str})
    df["mes_Proc"] = df["mes_Proc"].astype(str).str.strip()
    return df

def guardar_params(df: pd.DataFrame):
    wb = load_workbook(ARCHIVO_PARAMS)
    ws = wb[HOJA_PARAMS]
    ws.delete_rows(2, ws.max_row)
    for _, row in df.iterrows():
        ws.append(list(row))
    wb.save(ARCHIVO_PARAMS)

def render_parametros():
    st.markdown('<div class="param-title">📅 Gestión de parámetros mensuales</div>', unsafe_allow_html=True)
    if not os.path.exists(ARCHIVO_PARAMS):
        st.error(f"⚠️ No se encontró el archivo `{ARCHIVO_PARAMS}`.")
        return
    df_p = cargar_params()
    tab_add, tab_edit, tab_view = st.tabs(["➕ Agregar mes", "✏️ Editar mes", "📊 Ver tabla"])

    with tab_add:
        ultimo = df_p["mes_Proc"].dropna().iloc[-1] if not df_p.empty else "2026-01"
        try:
            dt_ult = datetime.strptime(str(ultimo)[:7], "%Y-%m")
            mes_sig = f"{dt_ult.year + 1}-01" if dt_ult.month == 12 else f"{dt_ult.year}-{dt_ult.month + 1:02d}"
        except Exception:
            mes_sig = ""
        nuevo_mes = st.text_input("Mes de proceso", value=mes_sig, placeholder="aaaa-mm", key="pm_nuevo_mes")
        mes_ok = False
        if nuevo_mes:
            try:
                dt = datetime.strptime(nuevo_mes[:7], "%Y-%m")
                ult_dia = calendar.monthrange(dt.year, dt.month)[1]
                mes_ok = True
                if nuevo_mes in df_p["mes_Proc"].values:
                    st.warning(f"⚠️ El mes **{nuevo_mes}** ya existe. Usa **Editar mes** para modificarlo.")
                    mes_ok = False
            except ValueError:
                st.error("Formato inválido. Usa aaaa-mm (ej: 2026-07)")
                ult_dia = 30
        if mes_ok:
            ult_row = df_p.iloc[-1] if not df_p.empty else {}
            def vref(c):
                try:
                    v = ult_row.get(c, 0)
                    return float(v) if pd.notna(v) else 0.0
                except Exception:
                    return 0.0
            nuevo = {"mes_Proc": nuevo_mes}
            campos = [c for c in LABELS_PARAMS if c not in ("mes_Proc", "Formato Fecha", "factor_sis", "ult_Diames")]
            cols_f = st.columns(3)
            for i, col in enumerate(campos):
                with cols_f[i % 3]:
                    fmt = "%.4f" if col in ("sis", "Aporte AFP", "Seg Social Exp vida", "aporte_Ccaf", "aporte_Fonasa") else "%.2f"
                    nuevo[col] = st.number_input(LABELS_PARAMS[col], value=vref(col), format=fmt, key=f"pm_new_{col}")
            nuevo["factor_sis"] = round(nuevo.get("sis", 0) / 100, 6)
            nuevo["ult_Diames"] = ult_dia
            nuevo["Formato Fecha"] = f"{ult_dia:02d}/{dt.month:02d}/{dt.year}"
            st.caption(f"📌 `factor_sis` = {nuevo['factor_sis']} | Último día del mes = {ult_dia}")
            if st.button("💾 Guardar nuevo mes", key="pm_btn_add"):
                fila = {col: nuevo.get(col) for col in df_p.columns}
                df_nuevo = pd.concat([df_p, pd.DataFrame([fila])], ignore_index=True)
                try:
                    guardar_params(df_nuevo)
                    st.cache_data.clear()
                    st.success(f"✅ Mes **{nuevo_mes}** agregado.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar: {e}")

    with tab_edit:
        meses = list(reversed(df_p["mes_Proc"].dropna().tolist()))
        mes_sel = st.selectbox("Mes a editar", meses, key="pm_mes_sel")
        if mes_sel:
            idx = df_p[df_p["mes_Proc"] == mes_sel].index[0]
            fila = df_p.loc[idx].copy()
            editado = {"mes_Proc": mes_sel}
            campos_e = [c for c in LABELS_PARAMS if c not in ("mes_Proc", "Formato Fecha", "factor_sis", "ult_Diames")]
            cols_e = st.columns(3)
            for i, col in enumerate(campos_e):
                with cols_e[i % 3]:
                    try:
                        val = float(fila.get(col, 0)) if pd.notna(fila.get(col)) else 0.0
                    except Exception:
                        val = 0.0
                    fmt = "%.4f" if col in ("sis", "Aporte AFP", "Seg Social Exp vida", "aporte_Ccaf", "aporte_Fonasa") else "%.2f"
                    editado[col] = st.number_input(LABELS_PARAMS[col], value=val, format=fmt, key=f"pm_edit_{col}")
            editado["factor_sis"] = round(editado.get("sis", 0) / 100, 6)
            try:
                dt_e = datetime.strptime(mes_sel[:7], "%Y-%m")
                ult_dia_e = calendar.monthrange(dt_e.year, dt_e.month)[1]
                editado["ult_Diames"] = ult_dia_e
                editado["Formato Fecha"] = f"{ult_dia_e:02d}/{dt_e.month:02d}/{dt_e.year}"
            except Exception:
                pass
            st.caption(f"📌 `factor_sis` = {editado['factor_sis']}")
            if st.button("💾 Guardar cambios", key="pm_btn_edit"):
                for col in df_p.columns:
                    df_p.at[idx, col] = editado.get(col, df_p.at[idx, col])
                try:
                    guardar_params(df_p)
                    st.cache_data.clear()
                    st.success(f"✅ Mes **{mes_sel}** actualizado.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Error al guardar: {e}")

    with tab_view:
        col_f, _ = st.columns([1, 3])
        with col_f:
            filtro = st.text_input("🔍 Filtrar por año", placeholder="ej: 2026", key="pm_filtro")
        df_v = df_p[df_p["mes_Proc"].str.startswith(filtro)] if filtro else df_p.copy()
        st.dataframe(df_v.rename(columns=LABELS_PARAMS), use_container_width=True, hide_index=True, height=480)
        st.caption(f"Total: {len(df_v)} períodos")

# ─────────────────────────────────────────────
# FUNCIONES DE PROCESAMIENTO
# ─────────────────────────────────────────────
def extraer_fecha_proceso(nombre_archivo):
    """Extrae los últimos 6 chars antes de .csv → formato aaaa-mm"""
    base = os.path.splitext(nombre_archivo)[0]
    sufijo = base[-6:]
    return f"{sufijo[:4]}-{sufijo[4:]}"

def safe_sum(df, cols):
    """Suma columnas que existen en el df, ignorando las que no existen."""
    cols_presentes = [c for c in cols if c in df.columns]
    if not cols_presentes:
        return pd.Series(0, index=df.index)
    return df[cols_presentes].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)

def get_col(df, col, default=0):
    """Obtiene una columna del df o retorna default si no existe."""
    if col in df.columns:
        return df[col].fillna(default)
    return pd.Series(default, index=df.index)

def validar_estructura(df, columnas_requeridas):
    """
    Verifica que un DataFrame tenga las columnas requeridas, sin importar
    el nombre del archivo subido. Retorna (ok, columnas_faltantes).
    """
    if df is None or df.empty:
        return False, columnas_requeridas
    faltantes = [c for c in columnas_requeridas if c not in df.columns]
    return len(faltantes) == 0, faltantes


def validar_archivos(archivos_subidos):
    """Valida que todos los archivos sean de la misma empresa.
    Para formato LRE (nombre empieza con dígito) compara prefijo de 10 chars.
    Para formato Rex+ (nombre libre) omite la validación por nombre."""
    nombres = [f.name for f in archivos_subidos]
    prefijos = [n[:10] for n in nombres]
    # Solo aplica si el nombre sigue convención RUT empresa (empieza con dígito)
    if not nombres[0][0].isdigit():
        return True, prefijos
    if len(set(prefijos)) > 1:
        return False, prefijos
    return True, prefijos

def detectar_formato_rexplus(df):
    """Retorna True si el CSV es formato Rex+ (tiene 'Id empleado' en lugar de 'Rut trabajador (1101)')."""
    return "Id empleado" in df.columns and "Rut trabajador (1101)" not in df.columns

def normalizar_rexplus(df):
    """
    Normaliza un DataFrame formato Rex+ al formato LRE que espera el programa.
    Retorna (df_normalizado, df_listado_empleados).
    """
    df = df.copy()

    # Convertir separador decimal de coma a punto en columnas numéricas
    for col in df.select_dtypes(include="object").columns:
        convertida = df[col].astype(str).str.replace(",", ".", regex=False)
        convertida_num = pd.to_numeric(convertida, errors="coerce")
        # Solo reemplazar si la mayoría de valores son numéricos
        if convertida_num.notna().sum() > df[col].notna().sum() * 0.5:
            df[col] = convertida_num

    # Combinar AFC empleador solidario + individual → AFC - Aporte empleador(4151)
    col_sol = "AFC - Aporte empleador solidario"
    col_ind = "AFC - Aporte empleador individual"
    if col_sol in df.columns or col_ind in df.columns:
        df["AFC - Aporte empleador(4151)"] = (
            pd.to_numeric(df.get(col_sol, 0), errors="coerce").fillna(0) +
            pd.to_numeric(df.get(col_ind, 0), errors="coerce").fillna(0)
        )
        df.drop(columns=[c for c in [col_sol, col_ind] if c in df.columns], inplace=True)

    # Construir listado_empleados antes de renombrar columnas
    df_empl = pd.DataFrame({
        "Rut":     df["Id empleado"].astype(str),
        "AFP":     df["afp"].astype(str) if "afp" in df.columns else "",
        "Isapre":  df["isapre"].astype(str) if "isapre" in df.columns else "",
        "Empresa": df["Id de empresa"].astype(str) if "Id de empresa" in df.columns else "",
    })

    # Renombrar columnas al estándar LRE
    df.rename(columns=COLS_REXPLUS_TO_LRE, inplace=True)

    return df, df_empl

def calcular_totales(df):
    """Calcula los 5 totales pre-validación."""
    df = df.copy()
    df["_total_haberes_afectos"] = safe_sum(df, COLS_HABERES_AFECTOS)
    df["_total_haberes_exentos"] = safe_sum(df, COLS_HABERES_EXENTOS)
    df["_total_descuentos_legales"] = safe_sum(df, COLS_DESCUENTOS_LEGALES)
    df["_total_otros_descuentos"] = safe_sum(df, COLS_OTROS_DESCUENTOS)
    df["_total_aportes_empleador"] = safe_sum(df, COLS_APORTES_EMPLEADOR)
    return df

def validar_cuadraturas(df, nombre_archivo):
    """Ejecuta las 6 validaciones y retorna lista de errores."""
    errores = []
    tol = 1  # tolerancia de 1 peso por redondeo

    validaciones = [
        ("V1", "_total_haberes_afectos", "Total haberes imponibles y tributables(5210)",
         "total_haberes_afectos ≠ Total haberes imponibles y tributables(5210)"),
        ("V2", "_total_haberes_exentos", "Total haberes no imponibles y no tributables(5230)",
         "total_haberes_exentos ≠ Total haberes no imponibles y no tributables(5230)"),
        ("V3", "_total_descuentos_legales", "Total descuentos por cotizaciones del trabajador(5341)",
         "total_descuentos_legales ≠ Total descuentos por cotizaciones del trabajador(5341)"),
        ("V4", "_total_otros_descuentos", "Total otros descuentos(5302)",
         "total_otros_descuentos ≠ Total otros descuentos(5302)"),
        ("V5", "_total_aportes_empleador", "Total aportes empleador(5410)",
         "total_aportes_empleador ≠ Total aportes empleador(5410)"),
    ]

    for codigo, col_calc, col_ctrl, mensaje in validaciones:
        if col_ctrl not in df.columns:
            continue
        ctrl = df[col_ctrl].fillna(0)
        calc = df[col_calc].fillna(0)
        mask = (calc - ctrl).abs() > tol
        filas_error = df[mask]
        for _, row in filas_error.iterrows():
            errores.append({
                "Archivo": nombre_archivo,
                "RUT": row.get("Rut trabajador (1101)", "N/D"),
                "Fila en archivo": int(row.get("_fila_csv", row.name + 2)),
                "Validación": codigo,
                "Descripción": mensaje,
                "Valor calculado": round(calc[row.name], 2),
                "Valor control": round(ctrl[row.name], 2),
                "Diferencia": round(calc[row.name] - ctrl[row.name], 2)
            })

    # V6: liquidez
    if all(c in df.columns for c in ["Total líquido(5501)"]):
        liq_calc = (df["_total_haberes_afectos"] + df["_total_haberes_exentos"]) - \
                   (df["_total_descuentos_legales"] + df["_total_otros_descuentos"])
        liq_ctrl = df["Total líquido(5501)"].fillna(0)
        mask = (liq_calc - liq_ctrl).abs() > tol
        filas_error = df[mask]
        for _, row in filas_error.iterrows():
            errores.append({
                "Archivo": nombre_archivo,
                "RUT": row.get("Rut trabajador (1101)", "N/D"),
                "Fila en archivo": int(row.get("_fila_csv", row.name + 2)),
                "Validación": "V6",
                "Descripción": "(haberes_afectos + haberes_exentos) - (desc_legales + otros_desc) ≠ Total líquido(5501)",
                "Valor calculado": round(liq_calc[row.name], 2),
                "Valor control": round(liq_ctrl[row.name], 2),
                "Diferencia": round(liq_calc[row.name] - liq_ctrl[row.name], 2)
            })

    return errores

def generar_filas_salida(df, fecha_proceso, refs):
    """Genera las filas del archivo de salida via pivot de conceptos."""
    filas = []

    # Cargar lookups
    equiv = refs.get("equiv_conceptos", pd.DataFrame())
    empleados = refs.get("listado_empleados", pd.DataFrame())
    empresas = refs.get("listado_empresas", pd.DataFrame())
    mutuales = refs.get("inst_mutuales", pd.DataFrame())
    cajas = refs.get("inst_cajas", pd.DataFrame())
    salud_inst = refs.get("inst_salud", pd.DataFrame())
    afp_inst   = refs.get("inst_afp", pd.DataFrame())

    def _norm(s):
        return str(s).lower().replace(" ","").replace("-","").replace("_","") if s else ""

    def _lookup_id(df, id_col, value):
        if df.empty or id_col not in df.columns or not value: return value
        match = df[df[id_col].apply(_norm) == _norm(value)]
        return match.iloc[0][id_col] if not match.empty else value

    cot_afp = refs.get("cot_afp_hist", pd.DataFrame())
    params = refs.get("parametros", pd.DataFrame())

    # Diccionario de equivalencias de conceptos
    equiv_dict = {}
    if not equiv.empty and "cod_lre" in equiv.columns and "concepto_detalle" in equiv.columns:
        equiv_dict = dict(zip(equiv["cod_lre"], equiv["concepto_detalle"]))

    # Parámetros mensuales — filtrar por mes de proceso
    tope_afp    = 0
    tope_ces    = 0
    tasa_sis    = 0
    tope_salud  = 0
    if not params.empty and "mes_Proc" in params.columns:
        fila_params = params[params["mes_Proc"].astype(str).str.strip() == str(fecha_proceso).strip()]
        if fila_params.empty:
            fila_params = params
        fp = fila_params.iloc[0]
        tope_afp   = pd.to_numeric(fp.get("topeImp_pesos_afp", 0), errors="coerce") or 0
        tope_ces   = pd.to_numeric(fp.get("topeCes_pesos",     0), errors="coerce") or 0
        tasa_sis   = pd.to_numeric(fp.get("sis",               0), errors="coerce") or 0
        tope_salud = pd.to_numeric(fp.get("topeSalud_pesos",   0), errors="coerce") or 0

    # Columnas de conceptos (las que están en equiv_dict)
    cols_concepto = [c for c in df.columns if c in equiv_dict]

    for _, row in df.iterrows():
        rut = row.get("Rut trabajador (1101)", "")

        # Lookup empresa
        _emp_raw = row.get("Id de empresa", "")
        _emp_num = pd.to_numeric(_emp_raw, errors="coerce")
        empresa_salida = str(int(_emp_num)) if not pd.isna(_emp_num) else str(_emp_raw or "").strip()
        if not empresa_salida and not empleados.empty and "Rut" in empleados.columns:
            emp_row = empleados[empleados["Rut"] == rut]
            if not emp_row.empty:
                nombre_empresa = emp_row.iloc[0].get("Empresa", "")
                if not empresas.empty and "Nombre" in empresas.columns:
                    emp2 = empresas[empresas["Nombre"] == nombre_empresa]
                    if not emp2.empty:
                        empresa_salida = emp2.iloc[0].iloc[0]

        # AFP e Isapre del empleado
        afp_empleado = ""
        isapre_empleado = ""
        if not empleados.empty and "Rut" in empleados.columns:
            emp_row = empleados[empleados["Rut"] == rut]
            if not emp_row.empty:
                afp_empleado = emp_row.iloc[0].get("AFP", "")
                isapre_empleado = emp_row.iloc[0].get("Isapre", "")

        dias_licencia = row.get("Nro días de licencia médica en el mes(1116)",
                         row.get("Nro días de licencia médica", None))
        _lic = pd.to_numeric(dias_licencia, errors="coerce"); dias_licencia = 0 if pd.isna(_lic) else int(_lic)
        _dias_trab_raw = row.get("Nro días trabajados en el mes(1115)",
                          row.get("Nro días trabajados", 0))
        _trab = pd.to_numeric(_dias_trab_raw, errors="coerce"); _dias_trab_raw = 0 if pd.isna(_trab) else int(_trab)
        dias_trabajados = max(0, _dias_trab_raw - dias_licencia)
        dias_vacaciones = row.get("Nro días de vacaciones en el mes(1117)", 0) or 0
        sueldo = row.get("Sueldo(2101)", 0) or 0
        total_imponible        = row.get("Total haberes imponibles y tributables(5210)", 0) or 0
        total_haberes_afectos  = row.get("_total_haberes_afectos", 0) or 0
        total_haberes_exentos  = row.get("_total_haberes_exentos", 0) or 0

        def _n(v):
            try: return float(str(v).replace(",", ".")) if pd.notna(v) and str(v).strip() != "" else 0.0
            except: return 0.0
        col_3143 = _n(row.get("Cotización obligatoria salud 7%(3143)", 0))
        col_3144 = _n(row.get("Cotización voluntaria para salud(3144)", 0))
        monto_isapre = col_3143 + col_3144
        col_3141 = _n(row.get("Cotización obligatoria previsional (AFP o IPS)(3141)", 0))
        col_3151 = _n(row.get("Cotización AFC - trabajador(3151)", 0))
        col_3154 = _n(row.get("Cotización adicional trabajo pesado - trabajador(3154)", 0))
        col_3156 = _n(row.get("Cotización APVi Mod B hasta UF50(3156)", 0))
        salud_afecto = min(col_3143 + col_3144, tope_salud) if tope_salud > 0 else col_3143 + col_3144
        rebaja_llss_impuesto = col_3141 + col_3151 + col_3154 + col_3156 + salud_afecto + _n(row.get("Rebaja zona extrema DL 889 (3167)", 0))
        col_1152 = row.get("Org. administrador ley 16.744(1152)", "")
        col_3110 = row.get("Crédito social CCAF(3110)", 0) or 0
        rebaja_zona = row.get("Rebaja zona extrema DL 889 (3167)", 0) or 0

        monto_init = (sueldo / dias_trabajados * 30) if dias_trabajados > 0 else 0

        # Fila por cada concepto
        CONCEPTOS_CON_CERO = {"impuesto", "cesEmpleado"}
        for col_csv in cols_concepto:
            _monto_raw = row.get(col_csv, 0)
            try:
                monto = float(str(_monto_raw).replace(",", ".")) if pd.notna(_monto_raw) and str(_monto_raw).strip() != "" else 0.0
            except (ValueError, TypeError):
                monto = 0.0
            id_concepto = equiv_dict.get(col_csv, "")

            if monto == 0 and id_concepto not in CONCEPTOS_CON_CERO:
                continue
            if id_concepto == "isapre":
                continue

            # Id de institución
            id_institucion = ""
            if id_concepto in GRUPOS_AFP:
                id_institucion = _lookup_id(afp_inst, "id_afp", afp_empleado)
            elif id_concepto == "apvi":
                _id_afp = _lookup_id(afp_inst, "id_afp", afp_empleado)
                id_institucion = f"apv{_id_afp}" if _id_afp else ""
            elif id_concepto in GRUPOS_ISAPRE:
                id_institucion = _lookup_id(salud_inst, "id_inst", isapre_empleado)
            elif id_concepto in GRUPOS_MUTUAL:
                if not mutuales.empty and "id_mutual" in mutuales.columns:
                    m = mutuales[mutuales["cod_lre"] == col_1152]
                    if m.empty: m = mutuales[mutuales["id_mutual"].apply(_norm) == _norm(col_1152)]
                    if m.empty and "nombre_mutual" in mutuales.columns:
                        m = mutuales[mutuales["nombre_mutual"].apply(_norm) == _norm(col_1152)]
                    if not m.empty: id_institucion = m.iloc[0]["id_mutual"]
            elif id_concepto in GRUPOS_CCAF and col_3110 != 0:
                if not cajas.empty:
                    c = cajas[cajas.iloc[:, 0] == col_3110]
                    if not c.empty and "id_ccaf" in cajas.columns:
                        id_institucion = c.iloc[0]["id_ccaf"]

            # Afecto
            afecto = ""
            if id_concepto in {"afp", "isapre", "mutual", "sis", "trabajoPesaEmpl"}:
                afecto = min(total_haberes_afectos, tope_afp) if tope_afp > 0 else total_haberes_afectos
            elif id_concepto in {"cesEmpleado", "cesAporteCi", "cesAporteSol"}:
                afecto = min(total_haberes_afectos, tope_ces) if tope_ces > 0 else total_haberes_afectos
            elif id_concepto in GRUPOS_AFP_MUTUAL_AFECTO:
                afecto = min(total_imponible, tope_afp) if tope_afp > 0 else total_imponible
            elif id_concepto in GRUPOS_CES_AFECTO:
                afecto = min(total_imponible, tope_ces) if tope_ces > 0 else total_imponible
            elif id_concepto == "totalesEmpl":
                afecto = total_haberes_afectos
            elif id_concepto == "impuesto":
                afecto = total_haberes_afectos - rebaja_llss_impuesto

            # Cotización de jubilación
            cot_jubilacion = 0
            if id_concepto == "afp":
                key_afp = f"{fecha_proceso}{id_institucion}"
                if not cot_afp.empty and "id_afp_hist" in cot_afp.columns:
                    r = cot_afp[cot_afp["id_afp_hist"] == key_afp]
                    if not r.empty:
                        cot_jubilacion = r.iloc[0].get("cot_hist_afp", 0) * 100
            elif id_concepto == "sis":
                key_sis = f"{fecha_proceso}{id_institucion}"
                if not cot_afp.empty and "id_afp_hist" in cot_afp.columns:
                    r = cot_afp[cot_afp["id_afp_hist"] == key_sis]
                    if not r.empty:
                        cot_jubilacion = r.iloc[0].get("sis_hist", 0) * 100
            elif id_concepto == "cesEmpleado":
                cot_jubilacion = 0.6
            elif id_concepto == "isapre":
                cot_jubilacion = monto
            elif id_concepto == "mutual":
                if not mutuales.empty and "cod_lre" in mutuales.columns and "nombre_mutual" in mutuales.columns:
                    m = mutuales[mutuales["cod_lre"] == col_1152]
                    if not m.empty:
                        nombre_mutual = m.iloc[0]["nombre_mutual"]
                        if not empresas.empty and "Mutual" in empresas.columns and "Cotización Mutual" in empresas.columns:
                            e = empresas[empresas["Mutual"] == nombre_mutual]
                            if not e.empty:
                                cot_jubilacion = e.iloc[0]["Cotización Mutual"]
            elif id_concepto == "licenciaDias":
                cot_jubilacion = ""
            elif id_concepto == "totalesEmpl":
                cot_jubilacion = min(total_haberes_afectos, tope_afp) if tope_afp > 0 else total_haberes_afectos

            filas.append({
                "Fecha de proceso": fecha_proceso,
                "Id empleado": rut,
                "Número de contrato": 1,
                "Id del concepto": id_concepto,
                "Monto del concepto": monto,
                "Afecto": afecto,
                "Id de institución": id_institucion,
                "Cotización de jubilación": cot_jubilacion,
                "Días de licencias": dias_licencia,
                "Días trabajados": dias_trabajados,
                "Fecha de aplicación": fecha_proceso,
                "Empresa": empresa_salida,
                "Total de rebajas por LLSS": rebaja_llss_impuesto if id_concepto == "impuesto" else 0,
                "Rentas no gravadas": total_haberes_exentos if id_concepto == "impuesto" else 0,
                "Rebaja por zona extrema": rebaja_zona if id_concepto == "impuesto" else 0,
                "Jornada": "C",
                "Días de vacaciones": dias_vacaciones,
                "Monto Init": monto_init if id_concepto == "sueldoBase" else 0,
                "Fase": 1,
            })

        # Fila isapre combinada
        if monto_isapre != 0:
            filas.append({
                "Fecha de proceso": fecha_proceso,
                "Id empleado": rut,
                "Número de contrato": 1,
                "Id del concepto": "isapre",
                "Monto del concepto": monto_isapre,
                "Afecto": min(total_haberes_afectos, tope_afp) if tope_afp > 0 else total_haberes_afectos,
                "Id de institución": _lookup_id(salud_inst, "id_inst", isapre_empleado),
                "Cotización de jubilación": monto_isapre,
                "Días de licencias": dias_licencia,
                "Días trabajados": dias_trabajados,
                "Fecha de aplicación": fecha_proceso,
                "Empresa": empresa_salida,
                "Total de rebajas por LLSS": 0,
                "Rentas no gravadas": 0,
                "Rebaja por zona extrema": 0,
                "Jornada": "C",
                "Días de vacaciones": dias_vacaciones,
                "Monto Init": 0,
                "Fase": 1,
            })

        # Fila adicional licenciaDias si aplica
        if dias_licencia > 0:
            filas.append({
                "Fecha de proceso": fecha_proceso,
                "Id empleado": rut,
                "Número de contrato": 1,
                "Id del concepto": "licenciaDias",
                "Monto del concepto": dias_licencia,
                "Afecto": "",
                "Id de institución": "",
                "Cotización de jubilación": "",
                "Días de licencias": dias_licencia,
                "Días trabajados": dias_trabajados,
                "Fecha de aplicación": fecha_proceso,
                "Empresa": empresa_salida,
                "Total de rebajas por LLSS": 0,
                "Rentas no gravadas": 0,
                "Rebaja por zona extrema": 0,
                "Jornada": "C",
                "Días de vacaciones": dias_vacaciones,
                "Monto Init": 0,
                "Fase": 1,
            })

    return pd.DataFrame(filas)

def generar_excel(df_salida):
    """Genera el Excel de salida con formato."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Migración"

    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="thin", color="E8EDF5"),
        right=Side(style="thin", color="E8EDF5")
    )

    cols = list(df_salida.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 14)

    for ri, row in enumerate(df_salida.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EAF0F8") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()

# ─────────────────────────────────────────────
def generar_log_excel(dfs):
    df = pd.concat(dfs, ignore_index=True)
    cols_int = [c for c in df.columns if c.startswith('_') and c not in
        ('_total_haberes_afectos','_total_haberes_exentos','_total_descuentos_legales','_total_otros_descuentos')]
    df = df.drop(columns=cols_int, errors='ignore')
    df = df.rename(columns={'_total_haberes_afectos':'Total haberes afectos',
        '_total_haberes_exentos':'Total haberes exentos',
        '_total_descuentos_legales':'Total descuentos legales',
        '_total_otros_descuentos':'Total otros descuentos'})
    df['Suma de haberes']   = df['Total haberes afectos'] + df['Total haberes exentos']
    df['Total descuentos']  = df['Total descuentos legales'] + df['Total otros descuentos']
    df['Liquido calculado'] = df['Suma de haberes'] - df['Total descuentos']
    col_liq = next((c for c in df.columns if '5501' in c), None)
    df['Diferencia'] = (df['Liquido calculado'] - pd.to_numeric(df[col_liq], errors='coerce').fillna(0)) if col_liq else 0
    out_buf = io.BytesIO()
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = 'Log validacion'
    hf=PatternFill('solid',fgColor='1A2744'); hfont=Font(bold=True,color='FFFFFF',size=10)
    yf=PatternFill('solid',fgColor='FFFF00'); ef=PatternFill('solid',fgColor='EAF0F8')
    wf=PatternFill('solid',fgColor='FFFFFF')
    brd=Border(bottom=Side(style='thin',color='E8EDF5'),right=Side(style='thin',color='E8EDF5'))
    for ci,col in enumerate(df.columns,1):
        cell=ws2.cell(row=1,column=ci,value=col); cell.fill=hf; cell.font=hfont
        cell.alignment=Alignment(horizontal='center',vertical='center')
        ws2.column_dimensions[cell.column_letter].width=max(len(str(col))+4,14)
    cols_list = list(df.columns)
    cols_amarillas = {"Total haberes afectos","Total haberes exentos","Total descuentos legales",
        "Total otros descuentos","Suma de haberes","Total descuentos","Liquido calculado","Diferencia"}
    idx_amarillas = {i+1 for i,c in enumerate(cols_list) if c in cols_amarillas}
    for ri,row in enumerate(df.itertuples(index=False),2):
        diff=row[-1]; row_diff=(isinstance(diff,(int,float)) and abs(diff)>1)
        base_fill = ef if ri%2==0 else wf
        for ci,val in enumerate(row,1):
            cell=ws2.cell(row=ri,column=ci,value=val)
            cell.fill = yf if (ci in idx_amarillas or row_diff) else base_fill
            cell.border=brd; cell.alignment=Alignment(vertical='center')
    ws2.freeze_panes='A2'; wb2.save(out_buf)
    return out_buf.getvalue()

# INTERFAZ PRINCIPAL
# ─────────────────────────────────────────────

# Header
st.markdown("""
<div class="rex-header">
    <div style="display:flex; align-items:center; gap:12px;">
        <div class="rex-logo">Rex<span>+</span></div>
        <span class="rex-title">Liquidaciones en detalle desde LRE</span>
    </div>
    <div class="rex-badge">PRODUCCIÓN</div>
</div>
""", unsafe_allow_html=True)

# ── NAVEGACIÓN PRINCIPAL ──
nav_dt, nav_migracion, nav_params = st.tabs([
    "🏛️ Migración archivo descargado desde DT",
    "📂 Migración desde archivo base LRE de Rex",
    "⚙️ Parámetros mensuales"
])

# Cargar referencias compartidas (disponibles para todos los tabs)
refs, errores_refs = cargar_referencias()
if errores_refs:
    st.markdown(f'<div class="alert-warning">⚠️ Archivos de referencia no encontrados en <b>/data</b>: {", ".join(errores_refs)}</div>', unsafe_allow_html=True)

# ── Carga opcional de equivalencia de conceptos ──
with st.expander("⚙️ Equivalencia de conceptos de la base", expanded=False):
    cargar_equiv_manual = st.checkbox(
        "¿Deseas cargar los conceptos de la base desde un archivo?",
        key="chk_equiv_conceptos",
        value=False,
        help="Activa esta opción para subir manualmente el archivo equiv_conceptos.xlsx."
    )
    if cargar_equiv_manual:
        col_eq, _ = st.columns([1, 1])
        with col_eq:
            archivo_equiv = st.file_uploader(
                "Sube el archivo equiv_conceptos.xlsx",
                type=["xlsx"],
                key="up_equiv_conceptos",
                help="Archivo con la equivalencia de conceptos LRE → concepto Rex+."
            )
        if archivo_equiv:
            refs["equiv_conceptos"] = pd.read_excel(archivo_equiv)
            st.markdown(
                f'<div class="alert-success">✅ Conceptos cargados desde archivo: <b>{archivo_equiv.name}</b> — {len(refs["equiv_conceptos"])} conceptos.</div>',
                unsafe_allow_html=True
            )
            mostrar_alerta_conceptos_prohibidos(verificar_conceptos_prohibidos(refs["equiv_conceptos"]))
        else:
            st.markdown(
                '<div class="alert-warning">⚠️ Sube el archivo <b>equiv_conceptos.xlsx</b> para usar una equivalencia personalizada.</div>',
                unsafe_allow_html=True
            )
    else:
        if "equiv_conceptos" in refs and not refs["equiv_conceptos"].empty:
            n = len(refs["equiv_conceptos"])
            st.markdown(
                f'<div class="alert-success">✅ Usando conceptos del servidor ({n} conceptos cargados).</div>',
                unsafe_allow_html=True
            )
            mostrar_alerta_conceptos_prohibidos(verificar_conceptos_prohibidos(refs["equiv_conceptos"]))

with nav_migracion:
    st.markdown('<div class="section-title">📂 Migración desde archivo base LRE de Rex</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Sube uno o más archivos CSV del mismo RUT empresa para generar el archivo de salida en Excel.</div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 1</div>
            <div class="step-title">Subir archivos CSV</div>
            <div class="step-desc">Uno o más archivos del mismo RUT empresa. Ej: 76247825-0_202601.csv</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 2</div>
            <div class="step-title">Validación automática</div>
            <div class="step-desc">Se verifican las cuadraturas contables de cada registro.</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 3</div>
            <div class="step-title">Descargar Excel</div>
            <div class="step-desc">Si todo cuadra, se genera el archivo de salida listo para importar.</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)

    # Uploaders lado a lado
    col_up1, col_up2 = st.columns(2)

    with col_up1:
        st.markdown("#### 📤 Subir archivos CSV")
        archivos = st.file_uploader(
            "Selecciona uno o más archivos CSV de Previred",
            type=["csv"],
            accept_multiple_files=True,
            help="Los archivos deben corresponder al mismo RUT empresa (primeros 10 caracteres del nombre)"
        )

    with col_up2:
        st.markdown("#### 👥 Listado de empleados del período")
        archivo_empleados = st.file_uploader(
            "Sube el archivo listado_empleados.xlsx correspondiente al período a procesar",
            type=["xlsx"],
            accept_multiple_files=False,
            help="Requerido solo para archivos LRE estándar. Debe contener: Rut, Empresa, AFP, Isapre. "
                 "No es necesario para archivos exportados directamente desde Rex+."
        )
        if archivo_empleados:
            st.markdown(f'<div class="alert-success">✅ Listado de empleados cargado: <b>{archivo_empleados.name}</b></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-warning">⚠️ Requerido solo para archivos LRE estándar. Los archivos exportados desde Rex+ no lo necesitan.</div>', unsafe_allow_html=True)

    col_up3, col_up4 = st.columns(2)
    with col_up3:
        st.markdown("#### 🏢 Listado de empresas del período")
        archivo_empresas_lre = st.file_uploader(
            "Sube el archivo listado_empresas.xlsx del período",
            type=["xlsx"],
            accept_multiple_files=False,
            key="up_empresas_lre",
            help="Requerido para resolver empresa e institución mutual."
        )
        if archivo_empresas_lre:
            st.markdown(f'<div class="alert-success">✅ Listado de empresas cargado: <b>{archivo_empresas_lre.name}</b></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-warning">⚠️ Requerido para resolver empresa e institución mutual.</div>', unsafe_allow_html=True)

    with col_up4:
        st.markdown("#### 📅 Parámetros mensuales")
        archivo_params = st.file_uploader(
            "Sube el archivo parametrosMesuales.xlsx",
            type=["xlsx"],
            accept_multiple_files=False,
            key="up_params",
            help="Si subes un archivo, actualizará automáticamente el servidor. Si no subes nada, se usará la versión guardada en el servidor."
        )
        if archivo_params:
            st.markdown(f'<div class="alert-success">✅ Parámetros cargados desde archivo: <b>{archivo_params.name}</b> — se actualizará el servidor.</div>', unsafe_allow_html=True)
        elif refs.get("parametros") is not None and not refs.get("parametros", pd.DataFrame()).empty:
            st.markdown('<div class="alert-success">✅ Usando parámetros del servidor. Puedes subir una versión actualizada aquí.</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-error">❌ No hay parámetros disponibles. Súbelos aquí o agrégalos en la pestaña <b>⚙️ Parámetros mensuales</b>.</div>', unsafe_allow_html=True)

    if archivos:
        st.markdown(f'<div class="alert-success">✅ {len(archivos)} archivo(s) cargado(s): {", ".join([f.name for f in archivos])}</div>', unsafe_allow_html=True)

        archivos_key = tuple(f.name for f in archivos)
        if st.session_state.get("_archivos_key") != archivos_key:
            st.session_state["_validacion_ok"] = False
            st.session_state["_archivos_key"] = archivos_key

        valido, prefijos = validar_archivos(archivos)
        if not valido:
            st.markdown(f"""
            <div class="alert-error">
                ❌ <b>Los archivos no corresponden a la misma empresa.</b><br>
                Se detectaron distintos RUT empresa: {", ".join(set(prefijos))}<br>
                Por favor sube solo archivos del mismo RUT empresa.
            </div>""", unsafe_allow_html=True)
            st.stop()

        if st.button("▶ Ejecutar validaciones"):
            todos_errores = []
            dfs = []
            df_empl_acum = pd.DataFrame()

            # ── Detectar formato leyendo el primer archivo ──
            primer_archivo = archivos[0]
            primer_archivo.seek(0)
            _ext0 = primer_archivo.name.lower().split(".")[-1]
            try:
                df_muestra = pd.read_csv(primer_archivo, encoding="utf-8-sig", sep=None, engine="python", nrows=1)
            except Exception:
                primer_archivo.seek(0)
                df_muestra = pd.read_csv(primer_archivo, encoding="latin-1", sep=None, engine="python", nrows=1)
            primer_archivo.seek(0)
            es_rexplus = detectar_formato_rexplus(df_muestra)

            if archivo_params:
                try:
                    df_params = pd.read_excel(archivo_params, sheet_name="Hoja2", dtype={"mes_Proc": str})
                except Exception:
                    archivo_params.seek(0)
                    df_params = pd.read_excel(archivo_params, sheet_name=0, dtype={"mes_Proc": str})
                if "mes_Proc" not in df_params.columns:
                    st.markdown(
                        '<div class="alert-error">❌ El archivo de parámetros no contiene la columna <b>mes_Proc</b>. '
                        'Verifica que estás subiendo <b>parametrosMesuales.xlsx</b> y que la hoja correcta tiene esa columna.</div>',
                        unsafe_allow_html=True
                    )
                    st.stop()
                df_params["mes_Proc"] = df_params["mes_Proc"].astype(str).str.strip()
                refs["parametros"] = df_params
                # Auto-guardar en data/ para que quede disponible en futuras sesiones
                try:
                    archivo_params.seek(0)
                    ruta_dest = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
                    with open(ruta_dest, "wb") as f_dest:
                        f_dest.write(archivo_params.read())
                    st.cache_data.clear()
                    st.markdown('<div class="alert-success">✅ Parámetros guardados en el servidor — disponibles para próximas sesiones.</div>', unsafe_allow_html=True)
                except Exception as e:
                    st.markdown(f'<div class="alert-warning">⚠️ No se pudo guardar en servidor: {e}</div>', unsafe_allow_html=True)
            elif refs.get("parametros") is not None and not refs.get("parametros", pd.DataFrame()).empty:
                df_params = refs["parametros"]
                # refs["parametros"] ya está cargado desde data/
            else:
                st.markdown(
                    '<div class="alert-error">❌ No hay parámetros disponibles. Súbelos aquí o agrégalos en la pestaña <b>⚙️ Parámetros mensuales</b>.</div>',
                    unsafe_allow_html=True
                )
                st.stop()

            # Cargar listado de empresas si fue subido
            if archivo_empresas_lre:
                df_empresas_lre = pd.read_excel(archivo_empresas_lre, header=1)
                df_empresas_lre.columns = [str(c).strip() for c in df_empresas_lre.columns]
                refs["listado_empresas"] = df_empresas_lre

            if es_rexplus:
                st.markdown('<div class="alert-success">✅ Formato detectado: <b>Rex+</b>. Los datos de AFP, Isapre y Mutual se obtienen del propio archivo.</div>', unsafe_allow_html=True)
            else:
                # Formato LRE estándar: requiere listado_empleados.xlsx
                if not archivo_empleados:
                    st.markdown('<div class="alert-error">❌ Debes subir el listado de empleados del período para archivos en formato LRE estándar.</div>', unsafe_allow_html=True)
                    st.stop()
                df_empleados_previred = pd.read_excel(archivo_empleados, header=1)
                df_empleados_previred.columns = [str(c).strip() for c in df_empleados_previred.columns]
                ok_emp, faltantes_emp = validar_estructura(df_empleados_previred, ["Rut", "Empresa", "AFP", "Isapre"])
                if not ok_emp:
                    st.markdown(f'<div class="alert-error">❌ <b>{archivo_empleados.name}</b>: Archivo no tiene la estructura esperada, corrija antes de continuar.</div>', unsafe_allow_html=True)
                    st.stop()
                refs["listado_empleados"] = df_empleados_previred

            barra_val = st.progress(0, text="Iniciando validaciones...")
            n_archivos = len(archivos)
            for i, archivo in enumerate(archivos):
                barra_val.progress(int((i / n_archivos) * 85) + 5, text=f"Leyendo {archivo.name}...")
                for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                    try:
                        archivo.seek(0)
                        df = pd.read_csv(archivo, encoding=enc, sep=None, engine="python")
                        df["_fila_csv"] = range(2, len(df) + 2)
                        break
                    except (UnicodeDecodeError, Exception):
                        continue
                else:
                    barra_val.empty()
                    st.error(f"❌ No se pudo leer {archivo.name}. Verifica que sea un CSV válido.")
                    st.stop()

                # ── Normalizar si es formato Rex+ ──
                if es_rexplus:
                    df, df_empl = normalizar_rexplus(df)
                    df_empl_acum = pd.concat([df_empl_acum, df_empl], ignore_index=True).drop_duplicates(subset=["Rut"])
                    refs["listado_empleados"] = df_empl_acum

                # ── Validar estructura del CSV ──
                if df.empty:
                    barra_val.empty()
                    st.markdown(f'<div class="alert-error">❌ <b>{archivo.name}</b>: Archivo no tiene la estructura esperada, corrija antes de continuar.</div>', unsafe_allow_html=True)
                    st.stop()
                if "Rut trabajador (1101)" not in df.columns:
                    barra_val.empty()
                    st.markdown(f'<div class="alert-error">❌ <b>{archivo.name}</b>: Archivo no tiene la estructura esperada, corrija antes de continuar.</div>', unsafe_allow_html=True)
                    st.stop()

                df = calcular_totales(df)
                errores = validar_cuadraturas(df, archivo.name)
                todos_errores.extend(errores)

                # Obtener fecha de proceso
                if es_rexplus and "Fecha de proceso" in df.columns:
                    df["_fecha_proceso"] = df["Fecha de proceso"].astype(str).str[:7].str.strip()
                else:
                    fecha_proceso = extraer_fecha_proceso(archivo.name)

                dfs.append(df)
            barra_val.progress(100, text="✅ Validaciones completadas")

            st.session_state["_validacion_ok"]  = True
            st.session_state["_todos_errores"]  = todos_errores
            st.session_state["_dfs"]            = dfs
            st.session_state["_refs_empl"]      = refs.get("listado_empleados", pd.DataFrame())
            st.session_state["_refs_params"]    = refs.get("parametros", pd.DataFrame())
            st.session_state["_nombre_empresa"] = archivos[0].name[:10]
            import os as _os
            st.session_state["_nombre_archivo"] = _os.path.splitext(archivos[0].name)[0]

        if st.session_state.get("_validacion_ok"):
            _todos_errores  = st.session_state["_todos_errores"]
            _dfs            = st.session_state["_dfs"]
            _refs_empl      = st.session_state["_refs_empl"]
            _nombre_empresa = st.session_state["_nombre_empresa"]
            _nombre_archivo = st.session_state.get("_nombre_archivo", _nombre_empresa)

            st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)
            st.markdown("### 🔍 Resultado de validaciones")

            if _todos_errores:
                st.markdown(f"""
                <div class="alert-error">
                    ❌ <b>No se puede generar el archivo de salida.</b><br>
                    Se encontraron <b>{len(_todos_errores)} error(es)</b> de validación en los registros procesados.
                </div>""", unsafe_allow_html=True)

                with st.expander("📋 Ver log de errores detallado"):
                    df_errores = pd.DataFrame(_todos_errores)
                    st.dataframe(df_errores, use_container_width=True, hide_index=True)
                    xlsx_log = generar_log_excel(_dfs)
                    st.download_button(
                        label="⬇️ Descargar log de errores (.xlsx)",
                        data=xlsx_log,
                        file_name=f"Log_{_nombre_archivo}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                st.markdown("""
                <div class="alert-success">
                    ✅ <b>Todas las validaciones se cumplieron correctamente.</b><br>
                    Los registros de todos los archivos cuadran sin diferencias.
                </div>""", unsafe_allow_html=True)

                # ── Configuración de Fase (destacada visualmente) ──
                st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)
                st.markdown('''<div style="background:#EFF6FF; border:2px solid #3B82F6; border-radius:8px; padding:14px 20px; margin-bottom:12px;">
<b style="font-size:1rem; color:#1D4ED8;">⚙️ Configuración de Fase</b><br>
<span style="color:#374151; font-size:0.9rem;">Si el cliente trabaja con el campo <b>Fase</b> en Rex+, actívalo antes de generar el archivo.</span>
</div>''', unsafe_allow_html=True)
                col_fase_chk, col_fase_num, _ = st.columns([1, 1, 2])
                with col_fase_chk:
                    usa_fase = st.checkbox("**¿El cliente usa Fase?**", key="mig_usa_fase")
                if usa_fase:
                    with col_fase_num:
                        numero_fase = st.number_input(
                            "Número de Fase",
                            min_value=1,
                            step=1,
                            value=1,
                            key="mig_numero_fase",
                            help="Ingresa el número de fase (entero mayor a 0)."
                        )
                else:
                    numero_fase = 0

                st.markdown("#### ¿Desea generar el archivo de salida?")
                col_a, col_b, col_c, _ = st.columns([1, 1, 1, 4])
                with col_a:
                    aceptar = st.button("✅ Aceptar")
                with col_b:
                    cancelar = st.button("✖ Cancelar")
                with col_c:
                    salir = st.button("🚪 Salir")

                if salir:
                    st.session_state["_validacion_ok"] = False
                    st.markdown('<div class="alert-warning">La sesión ha sido cerrada. Puedes cerrar esta ventana.</div>', unsafe_allow_html=True)
                    st.stop()
                if cancelar:
                    st.session_state["_validacion_ok"] = False
                    st.markdown('<div class="alert-warning">Operación cancelada. Puedes subir nuevos archivos.</div>', unsafe_allow_html=True)
                    st.stop()
                if aceptar:
                    refs["listado_empleados"] = _refs_empl
                    refs["parametros"] = st.session_state.get("_refs_params", refs.get("parametros", pd.DataFrame()))
                    barra_gen = st.progress(0, text="Generando registros de salida...")
                    df_combined = pd.concat(_dfs, ignore_index=True)
                    grupos = list(df_combined.groupby("_fecha_proceso"))
                    filas_salida = []
                    for gi, (_, grupo) in enumerate(grupos):
                        barra_gen.progress(int((gi / max(len(grupos), 1)) * 80) + 10, text=f"Procesando período {grupo['_fecha_proceso'].iloc[0]}...")
                        fp = grupo["_fecha_proceso"].iloc[0]
                        df_out = generar_filas_salida(grupo, fp, refs)
                        filas_salida.append(df_out)
                    df_final = pd.concat(filas_salida, ignore_index=True) if filas_salida else pd.DataFrame()
                    if usa_fase and numero_fase >= 1:
                        df_final["Fase"] = int(numero_fase)
                    else:
                        df_final = df_final.drop(columns=["Fase"], errors="ignore")
                    barra_gen.progress(95, text="Generando Excel...")
                    excel_bytes = generar_excel(df_final)
                    barra_gen.progress(100, text="✅ Archivo generado")
                    st.session_state["_validacion_ok"] = False
                    st.markdown('<div class="alert-success">✅ Archivo generado exitosamente.</div>', unsafe_allow_html=True)
                    st.download_button(
                        label="⬇️ Descargar archivo de salida (.xlsx)",
                        data=excel_bytes,
                        file_name=f"migracion_{_nombre_empresa}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

with nav_dt:
    render_modulo_dt(refs)

with nav_params:
    render_parametros()

