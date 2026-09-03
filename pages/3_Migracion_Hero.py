import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import io
import re
import unicodedata
from datetime import datetime
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
try:
    from modulo_dt import resolver_contrato as _resolver_contrato, cargar_empleados as _cargar_empleados
    _MODULO_DT_OK = True
except Exception:
    _MODULO_DT_OK = False

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
DATA_DIR = "data"
ARCHIVO_PARAMS  = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
ARCHIVO_EQUIV   = os.path.join(DATA_DIR, "equivalencias_libro_rex.xlsx")
ARCHIVO_EMPRESAS = os.path.join(DATA_DIR, "listado_empresas.xlsx")
ARCHIVO_AFP      = os.path.join(DATA_DIR, "inst_afp.xlsx")
ARCHIVO_SALUD    = os.path.join(DATA_DIR, "inst_salud.xlsx")
ARCHIVO_MUTUALES    = os.path.join(DATA_DIR, "inst_mutuales.xlsx")
ARCHIVO_COT_AFP_HIST= os.path.join(DATA_DIR, "cot_afp_hist.xlsx")

USA_FASES = False   # Cambiar a True si la empresa usa Fases en Rex+

st.set_page_config(
    page_title="Rex+ | Migración Hero",
    page_icon="📥",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.rex-header {
    background-color: #1a2744;
    padding: 14px 28px;
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 28px;
}
.rex-logo {
    background: white; color: #1a2744;
    font-weight: 800; font-size: 15px;
    padding: 5px 10px; border-radius: 6px; letter-spacing: 0.5px;
}
.rex-logo span { color: #00b4d8; }
.rex-title { color: white; font-size: 18px; font-weight: 600; margin-left: 16px; }
.rex-badge {
    background: #00b4d8; color: white;
    font-size: 11px; font-weight: 700;
    padding: 4px 12px; border-radius: 20px; letter-spacing: 1px;
}
.step-card {
    background: white; border: 1px solid #e8edf5;
    border-radius: 10px; padding: 18px 20px; margin-bottom: 12px;
}
.step-label { color: #00b4d8; font-size: 11px; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; margin-bottom: 4px; }
.step-title { font-size: 15px; font-weight: 600; color: #1a2744; margin-bottom: 4px; }
.step-desc  { font-size: 13px; color: #6b7a9a; }
.section-title { font-size: 20px; font-weight: 700; color: #1a2744; margin-bottom: 4px; }
.section-sub   { font-size: 13px; color: #6b7a9a; margin-bottom: 20px; }
.alert-error {
    background: #fff0f0; border-left: 4px solid #e53e3e;
    border-radius: 6px; padding: 12px 16px; margin: 8px 0; font-size: 13px; color: #c53030;
}
.alert-success {
    background: #f0fff4; border-left: 4px solid #38a169;
    border-radius: 6px; padding: 12px 16px; margin: 8px 0; font-size: 13px; color: #276749;
}
.alert-warning {
    background: #fffbf0; border-left: 4px solid #d69e2e;
    border-radius: 6px; padding: 12px 16px; margin: 8px 0; font-size: 13px; color: #744210;
}
.alert-info {
    background: #ebf8ff; border-left: 4px solid #3182ce;
    border-radius: 6px; padding: 12px 16px; margin: 8px 0; font-size: 13px; color: #2b6cb0;
}
.rex-divider { border: none; border-top: 1px solid #e8edf5; margin: 24px 0; }
.stButton > button {
    background-color: #1a2744 !important; color: white !important;
    border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; padding: 8px 20px !important; font-size: 14px !important;
}
.stButton > button:hover { background-color: #00b4d8 !important; color: white !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES — GRUPOS DE CONCEPTOS
# ─────────────────────────────────────────────
# Conceptos cuyo Afecto = Renta Imponible AFP del PDF
GRUPO_AFECTO_AFP = {
    "afp", "aporteAFPemp", "isapre", "mutual",
    "reliquidaAfp", "reliquidaIsapre", "reliquidaMutual",
    "trabajoPesa", "trabajoPesaEmpl", "voluntarioCoti", "voluntarioAhor",
    "reliquidaTrabEmpl", "reliquidaTrabPesa", "afpAhor",
}
# Afecto = Renta tributable del PDF
GRUPO_AFECTO_IMPUESTO = {"impuesto", "reliquidaImpuesto"}
# Afecto = suma haberes afectos del Libro
GRUPO_AFECTO_TOTALES  = {"totalesEmpl"}
# Afecto = min(suma haberes afectos, topeCes_pesos)
GRUPO_AFECTO_CES      = {"cesEmpleado", "reliquidaCesEmpl"}
# Afecto = afecto cesEmpleado (o cálculo histórico si hay licencia)
GRUPO_AFECTO_SIS      = {
    "sis", "aporteFAPPCEV", "cesAporteSol", "cesAporteCi",
    "reliquidaSis", "reliquidaCesSol", "reliquidaCesCi", "reliquidaAporteCEV",
}

# Conceptos con Id de institución
GRUPO_AFP_INST   = {"afp", "aporteAFPemp", "afpAhor", "voluntarioCoti", "voluntarioAhor",
                    "reliquidaAfp", "reliquidaAporteAFP", "trabajoPesa", "trabajoPesaEmpl",
                    "reliquidaTrabEmpl", "reliquidaTrabPesa", "sis", "reliquidaSis"}
GRUPO_ISAPRE_INST= {"isapre", "reliquidaIsapre"}
GRUPO_MUTUAL_INST= {"mutual", "reliquidaMutual"}
GRUPO_CES_INST   = {"cesEmpleado", "cesAporteSol", "cesAporteCi",
                    "reliquidaCesEmpl", "reliquidaCesSol", "reliquidaCesCi"}

# Conceptos con Cotización de jubilación (tasa AFP)
GRUPO_COT_JUBILACION = {"afp", "reliquidaAfp"}


# Parcial 7: solo para mutual/sis cuando hay licencia
GRUPO_PARCIAL7 = {"mutual", "sis", "reliquidaMutual", "reliquidaSis"}
# Parcial 8: solo para ces aportes
GRUPO_PARCIAL8 = {"cesAporteSol", "cesAporteCi", "reliquidaCesSol", "reliquidaCesCi"}

# Isapre: monto = suma de 3 columnas específicas del Libro
COLS_ISAPRE_LIBRO = [
    # Nombres reales en el Libro Peya/Stores
    "Isapre",
    "Isapre sobre 7%",
    "1524 COTIZACION FONASA",
    # aliases alternativos
    "1520 COTIZACION ISAPRE",
    "1521 COTIZACION ISAPRE ADICIONAL",
    "1524 Cotización FONASA", "1520 Cotizacion ISAPRE", "1521 Cotizacion ISAPRE Adicional",
]

# ─────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────
def _norm_col(s):
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)

def safe_num(val, default=0):
    """Convierte a float, retorna default si falla."""
    import math
    if isinstance(val, (int, float)):
        try:
            return default if math.isnan(float(val)) else float(val)
        except Exception:
            return default
    try:
        s = str(val).strip().replace("$", "").replace("\xa0", "").replace(" ", "")
        # Formato chileno: 1.234.567 (puntos = miles, coma = decimal)
        if s.count(".") > 1:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return default

def parse_monto_pdf(s):
    """Parsea monto del PDF: '3.569.575' → 3569575"""
    if not s:
        return 0
    s = str(s).strip().replace("$", "").strip()
    # Si tiene puntos como separadores de miles (formato chileno)
    partes = s.split(",")
    entero = partes[0].replace(".", "")
    decimal = partes[1] if len(partes) > 1 else "0"
    try:
        return float(f"{entero}.{decimal}")
    except Exception:
        return 0

def normalizar_rut(rut):
    """'10.540.597-9' → '105405979'"""
    if not rut:
        return ""
    return re.sub(r"[.\-]", "", str(rut)).strip().upper()

def formatear_rut(rut_norm):
    """'105405979' → '10540597-9'  (sin puntos, con guión)"""
    r = str(rut_norm).upper().replace(".", "").replace("-", "").strip()
    if len(r) < 2:
        return r
    return f"{r[:-1]}-{r[-1]}"

def safe_col_sum(row, columns):
    """Suma las columnas que existen en el row (pd.Series o dict)."""
    total = 0
    for c in columns:
        val = row.get(c, None) if isinstance(row, dict) else (row[c] if c in row.index else None)
        if val is not None and pd.notna(val):
            total += safe_num(val)
    return total

# ─────────────────────────────────────────────
# CARGA DE REFERENCIAS
# ─────────────────────────────────────────────
@st.cache_data
def cargar_equivalencias():
    if not os.path.exists(ARCHIVO_EQUIV):
        return pd.DataFrame()
    # header=2: las 2 primeras filas son título/subtítulo; los encabezados reales están en la fila 3
    df = pd.read_excel(ARCHIVO_EQUIV, header=2)
    df.columns = [str(c).strip() for c in df.columns]
    return df

@st.cache_data(ttl=0)
def cargar_params():
    if not os.path.exists(ARCHIVO_PARAMS):
        return pd.DataFrame()
    df = pd.read_excel(ARCHIVO_PARAMS, sheet_name=0, dtype={"mes_Proc": str})
    df["mes_Proc"] = df["mes_Proc"].astype(str).str.strip()
    return df

def get_params_mes(df_params, mes_proc):
    """Retorna dict con parámetros del mes (o último disponible)."""
    if df_params.empty:
        return {}
    fila = df_params[df_params["mes_Proc"].str.strip() == str(mes_proc).strip()]
    if fila.empty:
        fila = df_params
    f = fila.iloc[0]
    def _sf(v):
        try: return float(v)
        except: return 0
    return {k: (_sf(f[k]) if pd.notna(f.get(k, None)) else 0) for k in df_params.columns if k != "mes_Proc"}

def build_equiv_list(df_equiv):
    """Construye lista de mappings activos (con Concepto Rex+ no vacío)."""
    if df_equiv.empty:
        return []
    col_libro   = next((c for c in df_equiv.columns if "columna" in c.lower() and "libro" in c.lower()), None)
    col_concepto= next((c for c in df_equiv.columns if "concepto rex" in c.lower()), None)
    col_nombre  = next((c for c in df_equiv.columns if "nombre rex" in c.lower()), None)
    col_tipo    = next((c for c in df_equiv.columns if "tipo rex" in c.lower()), None)
    col_empresa = next((c for c in df_equiv.columns if c.lower() == "empresa"), None)

    if not all([col_libro, col_concepto]):
        return []

    result = []
    for _, row in df_equiv.iterrows():
        concepto = str(row.get(col_concepto, "") or "").strip()
        # Excluir filas sin concepto real
        if not concepto or concepto.lower() in ("nan", "") or concepto in ("—", "-", "(campo Afecto)"):
            continue
        cols_raw = str(row.get(col_libro, "") or "").strip()
        # Separadores: "/" "|" ";" — el Libro Talana usa "/"
        libro_cols = [c.strip() for c in re.split(r"[/|;]", cols_raw) if c.strip() and c.strip().lower() != "nan"]
        result.append({
            "concepto":   concepto,
            "nombre":     str(row.get(col_nombre, "") or "").strip(),
            "tipo":       str(row.get(col_tipo, "") or "").strip(),
            "empresa":    str(row.get(col_empresa, "") or "").strip(),
            "libro_cols": libro_cols,
        })
    return result

# ─────────────────────────────────────────────
# PARSEO DE PDF DE LIQUIDACIONES
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# LOOKUP AFP
# ─────────────────────────────────────────────
@st.cache_data
def cargar_inst_afp():
    """Devuelve dict: nombre_normalizado → id_afp."""
    try:
        df = pd.read_excel(ARCHIVO_AFP, dtype=str)
        mapping = {}
        for _, row in df.iterrows():
            nombre = str(row.get("nombre_afp", "") or "").strip().lower()
            id_afp = str(row.get("id_afp", "") or "").strip()
            if nombre and id_afp:
                # "afp provida" → "provida"
                clave = re.sub(r"^afp\s+", "", nombre).strip()
                mapping[clave] = id_afp
                mapping[nombre] = id_afp  # también con prefijo
        return mapping
    except Exception:
        return {}

def nombre_a_id_afp(nombre_pdf, inst_afp_dict):
    """Convierte nombre AFP del PDF (ej. 'Provida') → id_afp (ej. 'provida')."""
    if not nombre_pdf:
        return ""
    key = nombre_pdf.strip().lower()
    if key in inst_afp_dict:
        return inst_afp_dict[key]
    # sin prefijo "afp "
    key2 = re.sub(r"^afp\s+", "", key).strip()
    return inst_afp_dict.get(key2, nombre_pdf.lower())

@st.cache_data
def cargar_inst_salud():
    """Devuelve dict: nombre_normalizado → id_inst (isapre/fonasa)."""
    try:
        df = pd.read_excel(ARCHIVO_SALUD, dtype=str)
        mapping = {}
        for _, row in df.iterrows():
            nombre = str(row.get("nombre_inst", "") or "").strip().lower()
            id_i   = str(row.get("id_inst", "") or "").strip()
            if nombre and id_i:
                mapping[nombre] = id_i
                # sin prefijo "isapre "
                clave = re.sub(r"^isapre\s+", "", nombre).strip()
                mapping[clave] = id_i
        return mapping
    except Exception:
        return {}

def nombre_a_id_salud(nombre_pdf, inst_salud_dict):
    if not nombre_pdf:
        return ""
    key = nombre_pdf.strip().lower()
    if key in inst_salud_dict:
        return inst_salud_dict[key]
    key2 = re.sub(r"^isapre\s+", "", key).strip()
    return inst_salud_dict.get(key2, nombre_pdf.lower())

@st.cache_data
def cargar_inst_mutuales():
    """Devuelve dict: nombre_normalizado → id_mutual."""
    try:
        df = pd.read_excel(ARCHIVO_MUTUALES, dtype=str)
        mapping = {}
        for _, row in df.iterrows():
            nombre = str(row.get("nombre_mutual", "") or "").strip().lower()
            id_m   = str(row.get("id_mutual", "") or "").strip()
            if nombre and id_m:
                mapping[nombre] = id_m
                mapping[id_m]   = id_m  # el id se mapea a sí mismo
                # Alias por prefijos conocidos
                PREFIJOS = (
                    "asociacion chilena de seguridad", "achs",
                    "mutual de seguridad", "mutseg",
                    "instituto de seguridad del trabajo", "ist",
                    "sin mutual", "isl",
                )
                for alias in PREFIJOS:
                    if nombre.startswith(alias) or alias.startswith(nombre[:6]):
                        mapping[alias] = id_m
                # Alias adicionales según id
                _extra = {
                    "achs":   ["achs", "asociacion chilena", "a.c.h.s"],
                    "ist":    ["ist", "instituto de seguridad"],
                    "mutseg": ["mutual de seguridad", "mutual cchc", "mutseg"],
                    "isl":    ["isl", "sin mutual", "empresa aporta", "no tiene mutual"],
                }
                for aliases in _extra.get(id_m, []):
                    mapping[aliases] = id_m
        return mapping
    except Exception:
        return {}

@st.cache_data
def cargar_cot_afp_hist():
    """dict {id_afp_hist_lower: cot_hist_afp}. Clave = mes+id, ej: '2026-05provida'."""
    try:
        df = pd.read_excel(ARCHIVO_COT_AFP_HIST, dtype=str)
        return {
            str(r.get("id_afp_hist","") or "").strip().lower():
            str(r.get("cot_hist_afp","") or "").strip()
            for _, r in df.iterrows()
            if str(r.get("id_afp_hist","") or "").strip()
        }
    except Exception:
        return {}

def nombre_a_id_mutual(nombre_pdf, inst_mutuales_dict):
    if not nombre_pdf:
        return ""
    key = nombre_pdf.strip().lower()
    # 1. Si ya es un id válido (ej. "mutseg"), retornar directo
    if key in inst_mutuales_dict.values():
        return key
    # 2. Coincidencia exacta de clave
    if key in inst_mutuales_dict:
        return inst_mutuales_dict[key]
    # 3. Búsqueda parcial en ambas direcciones
    for k, v in inst_mutuales_dict.items():
        if k and (k in key or key in k):
            return v
    return nombre_pdf.lower()

def _match(pattern, text, group=1, flags=re.IGNORECASE):
    m = re.search(pattern, text, flags)
    return m.group(group).strip() if m else None

def parsear_pagina_pdf(text):
    """Parsea una página del PDF de liquidaciones. Retorna dict con datos del empleado."""
    emp = {}
    lines = text.split("\n")

    # RUT empleado (con puntos: "10.540.597-9")
    rut_raw = _match(r"RUT[\s:]+(\d{1,2}[\.\s]?\d{3}[\.\s]?\d{3}[\s]?-[\s]?[\dkK])", text)
    if rut_raw:
        emp["rut_display"] = rut_raw.replace(" ", "")
        emp["rut"] = normalizar_rut(rut_raw)

    # RUT empresa (sin puntos, sin espacio: "Rut:76211425-9")
    rut_emp_m = re.search(r"Rut:(\d{7,8}-[\dkK])", text)
    if rut_emp_m:
        emp["rut_empresa"] = rut_emp_m.group(1).strip()

    # Fecha ingreso TALANA: "Fecha de Ingreso: 31 de Agosto de 2020"
    fi_m = re.search(
        r"[Ff]echa\s+de\s+[Ii]ngreso[\s:]+(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})",
        text, re.IGNORECASE
    )
    if fi_m:
        _dia  = fi_m.group(1).zfill(2)
        _mes  = MESES_ES.get(fi_m.group(2).lower(), 0)
        _anio = fi_m.group(3)
        if _mes:
            emp["fecha_ingreso_ddmmaaaa"] = f"{_dia}-{str(_mes).zfill(2)}-{_anio}"
    else:
        # Fallback formato dd/mm/aaaa
        fi = _match(r"[Ff]echa\s+[Ii]ngreso[\s:]+(\d{2}/\d{2}/\d{4})", text)
        if fi:
            p = fi.split("/")
            emp["fecha_ingreso_ddmmaaaa"] = f"{p[0]}-{p[1]}-{p[2]}"

    # Días trabajados y licencia
    # TALANA: la fila de labels y la fila de valores son líneas separadas
    # Ejemplo:
    #   "Días Trabajados:    Días Licencia    Días Ausencia    Horas Base"
    #   "30    0    0    42,0"
    emp["dias_trabajados"] = 0
    emp["dias_licencia"]   = 0
    emp["horas_base"]      = 0
    for i, line in enumerate(lines):
        # Intento 1: label y valor en la misma línea (formato antiguo)
        if not emp["dias_trabajados"]:
            m = re.search(r"[Dd][ií]as?\s+[Tt]rab(?:ajados?)?[\s:]+([\d]+)", line)
            if m:
                emp["dias_trabajados"] = int(m.group(1))
        if not emp["dias_licencia"]:
            m = re.search(r"[Dd][ií]as?\s+[Ll]icencia[\s:]+([\d]+)", line)
            if m:
                emp["dias_licencia"] = int(m.group(1))
        if not emp["horas_base"]:
            m = re.search(r"[Hh]oras?\s+[Bb]ase[\s:]+([\d]+)", line)
            if m:
                emp["horas_base"] = int(m.group(1))
        # Intento 2: TALANA tabla — label en línea i, valores en línea i+1
        if re.search(r"[Dd][ií]as?\s+[Tt]rab", line) and re.search(r"[Dd][ií]as?\s+[Ll]icencia", line):
            # Línea siguiente contiene los valores numéricos separados por espacios
            next_line = lines[i + 1].strip() if i + 1 < len(lines) else ""
            nums = re.findall(r"(\d+)(?:[,.]\d+)?", next_line)
            if nums:
                # Orden en el PDF: Días Trab | Días Lic | Días Ausencia | Horas Base
                if not emp["dias_trabajados"] and len(nums) >= 1:
                    emp["dias_trabajados"] = int(nums[0])
                if not emp["dias_licencia"] and len(nums) >= 2:
                    emp["dias_licencia"] = int(nums[1])
                if not emp["horas_base"] and len(nums) >= 4:
                    emp["horas_base"] = int(nums[3])

    # Sueldo base (código 0001 o 1000)
    emp["sueldo_base"] = 0
    for line in lines:
        m = re.search(r"(?:0001|1000)\s+SUELDO\s+BASE\s+([\d.,]+)", line, re.IGNORECASE)
        if m:
            emp["sueldo_base"] = parse_monto_pdf(m.group(1))
            break

    # AFP: nombre, tasa, renta imponible
    emp["afp_nombre"]           = ""
    emp["afp_tasa"]             = ""
    emp["renta_imponible_afp"]  = 0
    for line in lines:
        # "1522 DESCUENTO AFP ( Provida: 11.45% : Renta Imponible: $ 3.569.575 )"
        m = re.search(
            r"1522\s+DESCUENTO\s+AFP\s*\(\s*(.+?)\s*:\s*([\d.,]+)\s*%\s*:\s*Renta\s+Imponible\s*:\s*\$?\s*([\d.,]+)",
            line, re.IGNORECASE
        )
        if m:
            emp["afp_nombre"]          = m.group(1).strip()
            emp["afp_tasa"]            = m.group(2).strip().replace(".", ",")  # formato "11,45"
            emp["renta_imponible_afp"] = parse_monto_pdf(m.group(3))
            break

    # ISAPRE / FONASA
    emp["salud_nombre"] = ""
    emp["salud_tipo"]   = ""
    for line in lines:
        m = re.search(r"1520\s+COTIZACION\s+ISAPRE\s*\(\s*([^)]+)\s*\)", line, re.IGNORECASE)
        if m:
            emp["salud_nombre"] = m.group(1).strip()
            emp["salud_tipo"]   = "ISAPRE"
            break
        m = re.search(r"1524\s+COTIZACION\s+FONASA", line, re.IGNORECASE)
        if m:
            emp["salud_nombre"] = "Fonasa"
            emp["salud_tipo"]   = "FONASA"
            break

    # Renta tributable (para concepto impuesto)
    emp["renta_tributable"] = 0
    for line in lines:
        m = re.search(r"[Rr]enta\s+[Tt]ributable[\s:$]+([\d.,]+)", line)
        if m:
            emp["renta_tributable"] = parse_monto_pdf(m.group(1))
            break

    # Total rebaja LLSS desde PDF (para concepto impuesto):
    # 1522 AFP + 1520 ISAPRE + 1523 Seg.Ces. + 1006 APV-B
    emp["rebaja_llss_pdf"]    = 0
    emp["total_exentos_pdf"]  = 0
    _rebaja = 0
    _exentos= 0
    CODIGOS_LLSS  = {"1522", "1520", "1523", "1006"}
    CODIGOS_EXENTO= {"1301", "1302", "1303", "1304", "1305", "1306", "1307", "1308",
                     "1309", "1310", "1311", "1312", "1313", "1314", "1315"}
    for line in lines:
        m = re.match(r"(\d{4})\s+.+?\s+([\d.,]+)\s*$", line.strip())
        if m:
            codigo = m.group(1)
            monto  = parse_monto_pdf(m.group(2))
            if codigo in CODIGOS_LLSS:
                _rebaja  += monto
            if codigo in CODIGOS_EXENTO:
                _exentos += monto
    emp["rebaja_llss_pdf"]   = _rebaja
    emp["total_exentos_pdf"] = _exentos

    # Período de la liquidación (Ej: "Período: Junio 2026" o "06/2026")
    emp["periodo_pdf"] = ""
    for pat in [
        r"[Pp]er[ií]odo[\s:Remuneraciones]+([A-Za-záéíóúÁÉÍÓÚ]+\s+\d{4})",
        r"[Pp]er[ií]odo[\s:]+(\d{2}/\d{4})",
        r"Liquidación\s+de\s+(?:Remuneraciones\s+)?([A-Za-záéíóú]+\s+\d{4})",
    ]:
        m_p = re.search(pat, text)
        if m_p:
            emp["periodo_pdf"] = m_p.group(1).strip()
            break

    return emp


def parsear_pdf_bytes(pdf_bytes):
    """Parsea todas las páginas del PDF. Retorna dict {rut_normalizado: emp_data}."""
    try:
        import pdfplumber
    except ImportError:
        return None, "pdfplumber no está instalado. Ejecuta: pip install pdfplumber"

    result = {}
    errores = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue
                try:
                    emp = parsear_pagina_pdf(text)
                    rut = emp.get("rut")
                    if rut:
                        result[rut] = emp
                except Exception as e:
                    errores.append(f"Pág {i+1}: {e}")
    except Exception as e:
        return None, str(e)

    return result, errores

def parsear_multiples_pdfs_historicos(pdf_files_list):
    """Parsea múltiples PDFs históricos. Retorna {rut_norm: [emp_data, ...]}."""
    hist = {}
    for f in pdf_files_list:
        pdf_dict, _ = parsear_pdf_bytes(f.read())
        if not pdf_dict:
            continue
        for rut, emp in pdf_dict.items():
            if rut not in hist:
                hist[rut] = []
            hist[rut].append(emp)
    return hist

MESES_ES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}

def periodo_pdf_a_aaaamm(periodo_str):
    """Convierte 'Junio 2026' o '06/2026' a '2026-06'. Retorna None si no puede."""""
    if not periodo_str:
        return None
    m = re.match(r"(\d{2})/(\d{4})", periodo_str.strip())
    if m:
        return f"{m.group(2)}-{m.group(1)}"
    m = re.match(r"([A-Za-záéíóúÁÉÍÓÚ]+)\s+(\d{4})", periodo_str.strip())
    if m:
        num = MESES_ES.get(m.group(1).lower())
        if num:
            return f"{m.group(2)}-{num:02d}"
    return None

def get_emp_sin_licencia(rut_norm, pdf_dict_historico):
    """Busca en el historial el primer mes sin licencia del empleado."""
    if not pdf_dict_historico or rut_norm not in pdf_dict_historico:
        return None
    for emp in pdf_dict_historico[rut_norm]:
        if emp.get("dias_licencia", 0) == 0:
            return emp
    return None

# ─────────────────────────────────────────────
# CARGA Y DETECCIÓN DEL LIBRO
# ─────────────────────────────────────────────
def cargar_libro(xlsx_bytes):
    """Carga el Libro de Remuneraciones. Auto-detecta fila de encabezados.
    Retorna (df, rut_empresa_libro) donde rut_empresa_libro se extrae de la fila 0."""
    # 1ra pasada: leer sin header para buscar la fila que contiene "RUT"
    raw = pd.read_excel(io.BytesIO(xlsx_bytes), header=None, dtype=str)
    header_row = 0
    # Extraer RUT empresa de fila 0
    rut_empresa_libro = ""
    if not raw.empty:
        fila0 = " ".join([str(v) for v in raw.iloc[0].values if pd.notna(v)])
        m_rut0 = re.search(r"(\d{7,8}-[\dkK])", fila0)
        if m_rut0:
            rut_empresa_libro = m_rut0.group(1)
    for i, row in raw.iterrows():
        vals = [str(v).strip().upper() for v in row.values if pd.notna(v)]
        if any("RUT" in v for v in vals):
            header_row = i
            break

    # 2da pasada: leer con el header correcto
    df = pd.read_excel(io.BytesIO(xlsx_bytes), header=header_row, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    # Eliminar filas completamente vacías
    df = df.dropna(how="all").reset_index(drop=True)
    # Convertir columnas numéricas
    for col in df.columns:
        try:
            converted = pd.to_numeric(df[col].str.replace(",", ".", regex=False), errors="coerce")
            if converted.notna().sum() > len(df) * 0.3:
                df[col] = converted
        except Exception:
            pass
    return df, rut_empresa_libro


def inferir_mes_desde_nombre(nombre_archivo):
    """Extrae aaaa-mm del nombre del archivo."""
    base = os.path.splitext(nombre_archivo)[0]
    meses = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
    }
    base_low = base.lower()
    for mes_nombre, mes_num in meses.items():
        if mes_nombre in base_low:
            m = re.search(r"(20\d{2})", base)
            if m:
                return f"{m.group(1)}-{mes_num}"
    # Intento numérico
    m = re.search(r"(20\d{2})[_\-]?(\d{2})", base)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return ""

# ─────────────────────────────────────────────
# LOOKUPS
# ─────────────────────────────────────────────
def _norm_str(s):
    return str(s or "").lower().strip().replace(" ", "").replace("-", "").replace("_", "")

def lookup_contrato_por_rut_inicio(rut_display, fecha_ingreso_ddmmaaaa, df_empleados):
    """Busca contrato en listado_empleados por clave RUT+FechaInicioContrato.
    Clave = RUT sin puntos (con guion) + fecha dd-mm-aaaa.
    Ej: '13552925-7' + '23-08-2021' → '13552925-723-08-2021'
    Columnas esperadas: 'Rut' y 'Fecha Inicio contrato'."""
    if df_empleados is None or df_empleados.empty or not rut_display or not fecha_ingreso_ddmmaaaa:
        return ""
    rut_pdf   = re.sub(r"\.", "", str(rut_display)).strip()   # "13552925-7"
    clave_pdf = rut_pdf + fecha_ingreso_ddmmaaaa               # "13552925-723-08-2021"

    col_rut = next((c for c in df_empleados.columns if c.strip().lower() == "rut"), None)
    col_fin = next((c for c in df_empleados.columns
                    if "fecha" in c.strip().lower() and "inicio" in c.strip().lower()), None)
    # Buscar columna "Contrato" exacta primero, luego parcial excluyendo "Fecha Inicio contrato"
    col_con = next((c for c in df_empleados.columns if c.strip().lower() == "contrato"), None)
    if not col_con:
        col_con = next((c for c in df_empleados.columns
                        if "contrato" in c.strip().lower()
                        and "fecha" not in c.strip().lower()
                        and "inicio" not in c.strip().lower()), None)
    if not col_rut or not col_fin or not col_con:
        return ""

    from datetime import datetime as _dt
    for _, row in df_empleados.iterrows():
        rut_emp = re.sub(r"\.", "", str(row.get(col_rut, "") or "")).strip()
        fi_val  = row.get(col_fin, "")
        if hasattr(fi_val, "strftime"):
            fi_str = fi_val.strftime("%d-%m-%Y")
        else:
            fi_str = str(fi_val).strip().split(" ")[0]  # recorta " 00:00:00" de fechas Excel
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    fi_str = _dt.strptime(fi_str, fmt).strftime("%d-%m-%Y")
                    break
                except Exception:
                    continue
        if (rut_emp + fi_str) == clave_pdf:
            return str(row.get(col_con, "")).strip()
    return ""


def lookup_contrato(rut_norm, fecha_proceso, df_empleados):
    """Busca contrato por RUT y mes de proceso en listado_empleados."""
    if df_empleados is None or df_empleados.empty:
        return ""
    if _MODULO_DT_OK:
        contrato, _, ok, _ = _resolver_contrato(df_empleados, rut_norm, fecha_proceso)
        return str(contrato) if ok else ""
    # Fallback sin modulo_dt
    col_rut = next((c for c in df_empleados.columns if "rut" in c.lower()), None)
    col_con = next((c for c in df_empleados.columns if "contrato" in c.lower()), None)
    if not col_rut or not col_con:
        return ""
    mask = df_empleados[col_rut].astype(str).str.upper().str.strip() == str(rut_norm).upper().strip()
    fila = df_empleados[mask]
    if fila.empty:
        return ""
    return str(fila.iloc[0][col_con])

def lookup_empresa_por_rut(rut_empresa, df_empresas):
    """Busca empresa en listado_empresas por columna 'Identificador nacional' (RUT empresa).
    Retorna el valor de la columna 'Empresa'."""
    if df_empresas is None or df_empresas.empty or not rut_empresa:
        return ""
    col_id  = next((c for c in df_empresas.columns
                    if "identificador" in c.lower() and "nacional" in c.lower()), None)
    col_emp = next((c for c in df_empresas.columns
                    if c.strip().lower() == "empresa"), None)
    if not col_id or not col_emp:
        return ""
    rut_q = re.sub(r"\.", "", str(rut_empresa)).strip().upper()
    for _, row in df_empresas.iterrows():
        rut_row = re.sub(r"\.", "", str(row.get(col_id, "") or "")).strip().upper()
        if rut_q == rut_row:
            return str(row.get(col_emp, "")).strip()
    return ""


def lookup_empresa_rex(empresa_nombre, df_empresas):
    """Dado el nombre de la empresa, retorna el código Empresa de Rex+ desde listado_empresas."""
    if df_empresas is None or df_empresas.empty:
        return ""
    col_emp = next((c for c in df_empresas.columns if "empresa" in c.lower()), None)
    if not col_emp:
        return ""
    # Buscar por nombre parcial
    norm_q = _norm_str(empresa_nombre)
    for _, row in df_empresas.iterrows():
        if norm_q in _norm_str(str(row.get(col_emp, ""))):
            return str(row[col_emp])
    # Si solo hay una empresa en el archivo, retornarla
    if len(df_empresas) == 1:
        return str(df_empresas.iloc[0][col_emp])
    return ""

def lookup_empresa_empleado(rut_norm, df_empleados):
    """Retorna el nombre de empresa del empleado."""
    if df_empleados is None or df_empleados.empty:
        return ""
    col_rut = next((c for c in df_empleados.columns if "rut" in c.lower()), None)
    col_emp = next((c for c in df_empleados.columns if "empresa" in c.lower()), None)
    if not col_rut or not col_emp:
        return ""
    mask = df_empleados[col_rut].apply(normalizar_rut) == rut_norm
    fila = df_empleados[mask]
    if fila.empty:
        return ""
    return str(fila.iloc[0][col_emp])

def get_mutual_nombre(rut_norm, df_empleados, df_empresas):
    """Retorna nombre de mutual del empleado según su empresa."""
    emp_nombre = lookup_empresa_empleado(rut_norm, df_empleados)
    if df_empresas is None or df_empresas.empty:
        return "Mutual de Seguridad"
    col_emp    = next((c for c in df_empresas.columns if "empresa" in c.lower()), None)
    col_mutual = next((c for c in df_empresas.columns if "mutual" in c.lower()), None)
    if not col_emp or not col_mutual:
        return "Mutual de Seguridad"
    norm_q = _norm_str(emp_nombre)
    for _, row in df_empresas.iterrows():
        if norm_q in _norm_str(str(row.get(col_emp, ""))):
            return str(row[col_mutual])
    if len(df_empresas) >= 1:
        return str(df_empresas.iloc[0].get(col_mutual, "Mutual de Seguridad"))
    return "Mutual de Seguridad"

def lookup_cotizacion_mutual(emp_nombre, df_empresas):
    """Trae 'Cotización Mutual' de la empresa desde listado_empresas."""
    if df_empresas is None or df_empresas.empty:
        return ""
    col_emp = next((c for c in df_empresas.columns if "empresa" in c.lower()), None)
    col_cot = next((c for c in df_empresas.columns
                    if "cot" in c.lower() and "mutual" in c.lower()), None)
    if not col_cot:
        mutuales_cols = [c for c in df_empresas.columns if "mutual" in c.lower()]
        col_cot = mutuales_cols[1] if len(mutuales_cols) > 1 else None
    if not col_emp or not col_cot:
        return ""
    norm_q = _norm_str(emp_nombre)
    for _, row in df_empresas.iterrows():
        if norm_q and norm_q in _norm_str(str(row.get(col_emp, ""))):
            return str(safe_num(row.get(col_cot, 0)))
    return str(safe_num(df_empresas.iloc[0].get(col_cot, 0))) if len(df_empresas) >= 1 else ""

# ─────────────────────────────────────────────
# CÁLCULO DE AFECTO
# ─────────────────────────────────────────────
def calcular_afecto(concepto, pdf_emp, suma_afectos_libro, params, cesEmpleado_afecto, pdf_emp_hist=None):
    """Calcula el campo Afecto según el concepto."""
    if concepto in GRUPO_AFECTO_AFP:
        return pdf_emp.get("renta_imponible_afp", 0)
    if concepto in GRUPO_AFECTO_IMPUESTO:
        return pdf_emp.get("renta_tributable", 0)
    if concepto in GRUPO_AFECTO_TOTALES:
        return suma_afectos_libro
    if concepto in GRUPO_AFECTO_CES:
        return cesEmpleado_afecto  # base CES derivada de 1523 SEGURO CESANTIA / 0.6%
    if concepto in GRUPO_AFECTO_SIS:
        dias_lic = pdf_emp.get("dias_licencia", 0)
        if dias_lic == 0:
            return cesEmpleado_afecto
        else:
            if pdf_emp_hist:
                renta_hist = pdf_emp_hist.get("renta_imponible_afp", 0)
                tope_ces = params.get("topeCes_pesos", 0)
                return min(renta_hist, tope_ces) if tope_ces else renta_hist
            return cesEmpleado_afecto
    return 0

def calcular_parcial7(concepto, params, dias_licencia=0, pdf_emp_hist=None):
    """Parcial 7: renta imponible del mes sin licencia (o tope fallback), para mutual/sis.
    dias_licencia se toma del libro Excel (columna Días Licencia)."""
    if concepto not in GRUPO_PARCIAL7:
        return 0
    if dias_licencia > 0:
        if pdf_emp_hist:
            renta_hist = pdf_emp_hist.get("renta_imponible_afp", 0)
            if renta_hist:
                return renta_hist
        # Fallback: tope AFP del mes
        return params.get("topeImp_pesos_afp", 0)
    return 0

def calcular_parcial8(concepto, cesEmpleado_afecto):
    """Parcial 8: afecto de cesEmpleado para cesAporte*."""
    if concepto in GRUPO_PARCIAL8:
        return cesEmpleado_afecto
    return 0

# ─────────────────────────────────────────────
# ETL PRINCIPAL
# ─────────────────────────────────────────────
def obtener_rut_libro(df_libro):
    """Detecta la columna RUT del trabajador en el Libro."""
    # Prioridad: columna que tenga "trabajador" o "empleado"
    for kw in ("trabajador", "empleado"):
        for c in df_libro.columns:
            if "rut" in c.lower() and kw in c.lower():
                return c
    # Fallback: primera columna con "rut"
    candidatos = [c for c in df_libro.columns if "rut" in c.lower()]
    if candidatos:
        return candidatos[0]
    # Buscar columna con formato de RUT
    for col in df_libro.columns[:5]:
        sample = df_libro[col].dropna().head(5).astype(str)
        if sample.str.match(r"\d{1,2}\.?\d{3}\.?\d{3}-[\dkK]").any():
            return col
    return None

def generar_archivo_salida(
    df_libro, pdf_dict, df_empleados, df_empresas, df_params,
    equiv_list, fecha_proceso, nombre_libro, usa_fases=False,
    pdf_dict_historico=None,
    rut_empresa_libro=""
):
    """ETL principal: Libro + PDF + maestros → DataFrame de salida Rex+."""
    params = get_params_mes(df_params, fecha_proceso)
    tope_ces  = params.get("topeCes_pesos", 0)
    tope_afp  = params.get("topeImp_pesos_afp", 0)

    col_rut = obtener_rut_libro(df_libro)
    if not col_rut:
        return None, "No se encontró columna RUT en el Libro."

    _col_norm_idx = {}
    for c in df_libro.columns:
        key = _norm_col(c)
        if key not in _col_norm_idx:
            _col_norm_idx[key] = c
        m_dup = re.match(r"^(.+)\.\d+$", c)
        if m_dup:
            key2 = _norm_col(m_dup.group(1))
            if key2 not in _col_norm_idx:
                _col_norm_idx[key2] = c

    # Columnas del libro para rebajas LLSS (concepto impuesto)
    # = suma de todos los haberes exentos del libro
    # Primero buscar columna resumen "Remuneración No Imponible"
    _col_no_imponible = next(
        (c for c in df_libro.columns if "no imponible" in c.lower() or
         ("remunera" in c.lower() and "exento" in c.lower())),
        None
    )
    # Códigos haberes exentos para suma por columnas individuales (fallback)
    _CODIGOS_EXENTOS_LIBRO = {
        "0234", "0232", "0233", "0247", "0094", "0095",
        "0241", "0243", "0245", "0268", "0253", "0269",
        "0231", "0902", "0907", "0908", "0910",
    }
    _cols_rebaja_llss = []
    if not _col_no_imponible:
        for _col in df_libro.columns:
            _col_l = _col.strip()
            # Columna comienza con un código de 4 dígitos que está en la lista
            m_cod = re.match(r"^0?(\d{3,4})", _col_l)
            if m_cod:
                cod4 = m_cod.group(0).zfill(4)
                if cod4 in _CODIGOS_EXENTOS_LIBRO:
                    _cols_rebaja_llss.append(_col)

    # Filtrar equivalencias válidas para esta empresa (según nombre archivo)
    empresa_libro = ""
    if "stores" in nombre_libro.lower():
        empresa_libro = "DH Stores"
    elif "peya" in nombre_libro.lower() or "ecommerce" in nombre_libro.lower() or "delivery hero ec" in nombre_libro.lower():
        empresa_libro = "DH E-Commerce"

    inst_afp_dict     = cargar_inst_afp()
    inst_salud_dict   = cargar_inst_salud()
    inst_mutuales_dict= cargar_inst_mutuales()
    cot_afp_hist_dict = cargar_cot_afp_hist()
    # Empresa global desde RUT fila 0 del libro
    empresa_code_global = lookup_empresa_por_rut(rut_empresa_libro, df_empresas) if rut_empresa_libro else ""

    equiv_activos = [
        m for m in equiv_list
        if m["concepto"] and m.get("tipo","").lower() != "dato" and (
            not empresa_libro or                                         # sin empresa detectada: aceptar todo
            m["empresa"].lower() in ("ambas", "ambas empresas", "") or
            empresa_libro.lower() in m["empresa"].lower()
        )
    ]

    filas = []
    warnings = []

    for _, libro_row in df_libro.iterrows():
        rut_raw  = str(libro_row.get(col_rut, "") or "")
        rut_norm = normalizar_rut(rut_raw)
        if not rut_norm or rut_norm.lower() in ("nan", ""):
            continue

        rut_fmt      = formatear_rut(rut_norm)
        pdf_emp      = pdf_dict.get(rut_norm, {})
        pdf_emp_hist = get_emp_sin_licencia(rut_norm, pdf_dict_historico)
        _rut_display = pdf_emp.get("rut_display", "")
        _fi_ddmm     = pdf_emp.get("fecha_ingreso_ddmmaaaa", "")
        contrato     = lookup_contrato_por_rut_inicio(_rut_display, _fi_ddmm, df_empleados)
        if not contrato:
            contrato = lookup_contrato(rut_norm, mes_proceso, df_empleados)

        # Empresa Rex+: primero desde RUT fila 0 del libro, luego PDF, luego nombre empleado
        empresa_code = empresa_code_global
        if not empresa_code:
            _rut_empresa = pdf_emp.get("rut_empresa", "")
            empresa_code = lookup_empresa_por_rut(_rut_empresa, df_empresas)
        if not empresa_code:
            emp_nombre   = lookup_empresa_empleado(rut_norm, df_empleados)
            empresa_code = lookup_empresa_rex(emp_nombre, df_empresas)
        emp_nombre   = lookup_empresa_empleado(rut_norm, df_empleados)
        mutual_nombre= get_mutual_nombre(rut_norm, df_empleados, df_empresas)

        # Días — leer desde el libro Excel; fallback al PDF
        dias_trab = pdf_emp.get("dias_trabajados", 0)
        horas_base= pdf_emp.get("horas_base", 0)
        # Columna "Días Licencia" del libro (búsqueda flexible)
        _col_dias_lic = next(
            (c for c in df_libro.columns
             if re.search(r"d[ií]as?\s+licencia", c, re.IGNORECASE)),
            None
        )
        if _col_dias_lic:
            dias_lic = int(safe_num(libro_row.get(_col_dias_lic, 0)))
        else:
            dias_lic = pdf_emp.get("dias_licencia", 0)

        # Jornada: horas_base > 30 → "C", sino "P"
        jornada   = "C" if horas_base > 30 else "P"

        # Monto Init = (sueldo_base / dias_trab) * 30
        sueldo_base_pdf = pdf_emp.get("sueldo_base", 0)
        if dias_trab and dias_trab > 0 and sueldo_base_pdf:
            monto_init = round((sueldo_base_pdf / dias_trab) * 30)
        else:
            monto_init = sueldo_base_pdf

        # Suma haberes afectos del Libro (para Afecto de totalesEmpl y cesEmpleado)
        # Buscar columna totales en el libro o suma por secciones
        suma_afectos = 0
        col_total_afectos = next((c for c in df_libro.columns if "total" in c.lower() and "afect" in c.lower()), None)
        if not col_total_afectos:
            col_total_afectos = next((c for c in df_libro.columns if "total haber" in c.lower() and "impon" in c.lower()), None)
        if not col_total_afectos:
            # Libro Talana: "Remuneración Imponible"
            col_total_afectos = next((c for c in df_libro.columns if "imponible" in c.lower() and "remunera" in c.lower()), None)
        if col_total_afectos:
            suma_afectos = safe_num(libro_row.get(col_total_afectos, 0))

        # cesEmpleado afecto: base CES = 1523 SEGURO CESANTIA / 0.6%
        # (tope CES es distinto y mayor al tope AFP usado en suma_afectos)
        _col_ces_empl = next(
            (c for c in df_libro.columns if "1523" in c and "cesantia" in c.lower()), None
        )
        _ces_empl_monto = safe_num(libro_row.get(_col_ces_empl, 0)) if _col_ces_empl else 0
        if _ces_empl_monto > 0:
            cesEmp_afecto = round(_ces_empl_monto / 0.006)
        else:
            cesEmp_afecto = min(suma_afectos, tope_ces) if tope_ces else suma_afectos

        # Calcular Parcial 8 base (necesita cesEmp_afecto)
        # Lo calculamos por empleado y lo reutilizamos

        _conceptos_ya_vistos = set()
        for mapping in equiv_activos:
            concepto  = mapping["concepto"]
            libro_cols= mapping["libro_cols"]
            # Deduplicar conceptos que usan COLS fijas (isapre, impuesto)
            # Para impuesto: usar solo la primera mapping (1527 IMPUESTO UNICO),
            # ignorar filas adicionales como 1529 IMPUESTO UNICO FINIQUITO
            if concepto in GRUPO_ISAPRE_INST or concepto == "impuesto":
                if concepto in _conceptos_ya_vistos:
                    continue
                _conceptos_ya_vistos.add(concepto)

            # ── Monto del concepto ──
            if concepto in GRUPO_ISAPRE_INST:
                # Isapre: suma de columnas FONASA + ISAPRE + ISAPRE ADICIONAL
                monto = safe_col_sum(libro_row, COLS_ISAPRE_LIBRO)
            else:
                monto = 0
                for lc in libro_cols:
                    col_real = lc if lc in df_libro.columns else _col_norm_idx.get(_norm_col(lc))
                    if col_real:
                        monto += safe_num(libro_row.get(col_real, 0))
                # Fallback impuesto: buscar 2ª ocurrencia 'Impuestos' o '1527 IMPUESTO UNICO'
                if concepto in GRUPO_AFECTO_IMPUESTO and monto == 0:
                    for _try_col in ["1527 IMPUESTO UNICO", "Impuestos.1", "Impuestos"]:
                        if _try_col in df_libro.columns:
                            monto = safe_num(libro_row.get(_try_col, 0))
                            if monto:
                                break

            # Omitir si monto = 0 (excepto conceptos resumen que siempre van)
            CONCEPTOS_SIEMPRE = {"totalesEmpl", "impuesto"}
            if monto == 0 and concepto not in CONCEPTOS_SIEMPRE:
                continue

            # ── Afecto ──
            afecto = calcular_afecto(concepto, pdf_emp, suma_afectos, params, cesEmp_afecto, pdf_emp_hist=pdf_emp_hist)

            # ── Id de institución ──
            _CES_AFP = {"cesEmpleado", "cesAporteCi", "cesAporteSol",
                        "reliquidaCesEmpl", "reliquidaCesSol", "reliquidaCesCi"}
            _FAPP    = {"aporteFAPPCEV", "reliquidaAporteCEV"}
            id_inst = ""
            if concepto in GRUPO_AFP_INST or concepto in _CES_AFP:
                id_inst = nombre_a_id_afp(pdf_emp.get("afp_nombre", ""), inst_afp_dict)
            elif concepto in _FAPP:
                id_inst = "seguridadsocial"
            elif concepto in GRUPO_ISAPRE_INST:
                id_inst = nombre_a_id_salud(pdf_emp.get("salud_nombre", ""), inst_salud_dict)
            elif concepto in GRUPO_MUTUAL_INST:
                id_inst = nombre_a_id_mutual(mutual_nombre, inst_mutuales_dict)

            # ── Cotización de Jubilación ──
            def _fmt_cot(v, entero=False):
                """Formatea valor para col. Cotización de Jubilación.
                Separador decimal: coma. entero=True → sin decimales."""
                if v == "" or v is None:
                    return ""
                try:
                    f = float(str(v).replace(",", "."))
                    if entero:
                        return str(int(round(f)))
                    s = f"{f:.10f}".rstrip("0").rstrip(".")
                    return s.replace(".", ",")
                except Exception:
                    return str(v)

            cot_jubilacion = ""
            if concepto == "afp":
                # Tasa histórica AFP: dict guarda valor %, se divide /100
                _id_afp_cj = nombre_a_id_afp(
                    pdf_emp.get("afp_nombre",""), inst_afp_dict).lower()
                _raw_afp = cot_afp_hist_dict.get(f"{fecha_proceso}{_id_afp_cj}", "")
                if _raw_afp:
                    try:
                        cot_jubilacion = _fmt_cot(float(str(_raw_afp).replace(',', '.')) * 100)
                    except Exception:
                        cot_jubilacion = _fmt_cot(_raw_afp)
            elif concepto == "mutual":
                cot_jubilacion = _fmt_cot(lookup_cotizacion_mutual(emp_nombre, df_empresas))
            elif concepto == "sis":
                cot_jubilacion = _fmt_cot(safe_num(params.get("sis", 0)))
            elif concepto == "totalesEmpl":
                cot_jubilacion = str(int(round(suma_afectos))) if suma_afectos else ""
            elif concepto == "aporteAFPemp":
                cot_jubilacion = _fmt_cot(safe_num(params.get("Aporte AFP", 0)))
            elif concepto == "aporteFAPPCEV":
                cot_jubilacion = _fmt_cot(safe_num(params.get("Seg Social Exp vida", 0)))
            elif concepto in GRUPO_ISAPRE_INST:
                cot_jubilacion = _fmt_cot(monto, entero=True)
            elif concepto == "cesEmpleado":
                cot_jubilacion = "0,6"

            # ── Total rebajas LLSS ──
            # Solo para impuesto: suma columnas descuentos legales del libro
            total_rebajas_llss = 0
            if concepto in GRUPO_AFECTO_IMPUESTO:
                _COLS_REBAJA_LLSS_FIJAS = [
                    "1006 APV REGIMEN B",
                    "1524 COTIZACION FONASA",
                    "Isapre",
                    "Isapre sobre 7%",
                    "1523 SEGURO CESANTIA",
                ]
                for _c in _COLS_REBAJA_LLSS_FIJAS:
                    col_real = _c if _c in df_libro.columns else _col_norm_idx.get(_norm_col(_c))
                    if col_real:
                        total_rebajas_llss += safe_num(libro_row.get(col_real, 0))

            # ── Rentas no gravadas ──
            # Solo aplica para concepto impuesto; resto = 0
            rentas_no_gravadas = 0
            if concepto in GRUPO_AFECTO_IMPUESTO:
                if _col_no_imponible:
                    rentas_no_gravadas = safe_num(libro_row.get(_col_no_imponible, 0))
                elif _cols_rebaja_llss:
                    rentas_no_gravadas = sum(safe_num(libro_row.get(c, 0)) for c in _cols_rebaja_llss)

            # ── Parcial 7 ──
            parcial7 = calcular_parcial7(concepto, params, dias_licencia=dias_lic, pdf_emp_hist=pdf_emp_hist)

            # ── Parcial 8 ──
            parcial8 = calcular_parcial8(concepto, cesEmp_afecto)

            # ── Fecha aplicación = fecha proceso ──
            fila = {
                "Fecha de proceso":         fecha_proceso,
                "Id empleado":              rut_fmt,
                "Número de contrato":       contrato,
                "Id del concepto":          concepto,
                "Monto del concepto":       int(round(monto)) if monto else 0,
                "Afecto":                   int(round(afecto)) if afecto else 0,
                "Id de institución":        id_inst,
                "Cotización de jubilación": cot_jubilacion,
                "Días de licencias":        dias_lic,
                "Días trabajados":          dias_trab,
                "Fecha de aplicación":      fecha_proceso,
                "Empresa":                  empresa_code,
                "Total de rebajas por LLSS": int(round(total_rebajas_llss)) if total_rebajas_llss else 0,
                "Rentas no gravadas":       int(round(rentas_no_gravadas)) if rentas_no_gravadas else 0,
                "Rebaja por zona extrema":  0,
                "Jornada":                  jornada,
                "Días de vacaciones":       0,
                "Monto Init":               int(round(monto_init)) if monto_init else 0,
                "Parcial 7":                int(round(parcial7)) if parcial7 else 0,
                "Parcial 8":                int(round(parcial8)) if parcial8 else 0,
            }
            if usa_fases:
                fila["Fase"] = 1

            filas.append(fila)

    if not filas:
        return pd.DataFrame(), warnings

    df_salida = pd.DataFrame(filas)
    return df_salida, warnings

# ─────────────────────────────────────────────
# GENERACIÓN DE EXCEL DE SALIDA
# ─────────────────────────────────────────────
def generar_excel_salida(df_salida):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Rex+ Importación"

    hdr_fill = PatternFill("solid", fgColor="1A2744")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    brd = Border(
        bottom=Side(style="thin", color="E8EDF5"),
        right =Side(style="thin", color="E8EDF5"),
    )

    cols = list(df_salida.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 14)

    for ri, row in enumerate(df_salida.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EAF0F8") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = brd
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()

# ─────────────────────────────────────────────
# INTERFAZ PRINCIPAL
# ─────────────────────────────────────────────
st.markdown("""
<div class="rex-header">
    <div style="display:flex; align-items:center; gap:12px;">
        <div class="rex-logo">Rex<span>+</span></div>
        <span class="rex-title">Libro de Remuneraciones → Importación detalle</span>
    </div>
    <div class="rex-badge">PRODUCCIÓN</div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="section-title">📥 Importación desde Libro de Remuneraciones</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Procesa el Libro Excel + liquidaciones PDF y genera el archivo de importación Rex+.</div>', unsafe_allow_html=True)

# Pasos
col1, col2, col3, col4 = st.columns(4)
with col1:
    st.markdown("""<div class="step-card">
        <div class="step-label">PASO 1</div>
        <div class="step-title">Libro xlsx</div>
        <div class="step-desc">Libro de Remuneraciones del período (Peya o Stores).</div>
    </div>""", unsafe_allow_html=True)
with col2:
    st.markdown("""<div class="step-card">
        <div class="step-label">PASO 2</div>
        <div class="step-title">PDF liquidaciones</div>
        <div class="step-desc">Liquidaciones de sueldo TALANA del mismo período.</div>
    </div>""", unsafe_allow_html=True)
with col3:
    st.markdown("""<div class="step-card">
        <div class="step-label">PASO 3</div>
        <div class="step-title">Archivos maestros</div>
        <div class="step-desc">Listado empleados, empresas y parámetros mensuales.</div>
    </div>""", unsafe_allow_html=True)
with col4:
    st.markdown("""<div class="step-card">
        <div class="step-label">PASO 4</div>
        <div class="step-title">Descargar Excel</div>
        <div class="step-desc">Archivo listo para importar en Rex+ Remuneraciones.</div>
    </div>""", unsafe_allow_html=True)

st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)

# ── CARGAR REFERENCIAS ──
refs_equiv  = cargar_equivalencias()
refs_params = cargar_params()

if refs_equiv.empty:
    st.markdown(
        f'<div class="alert-warning">⚠️ No se encontró <b>data/equivalencias_libro_rex.xlsx</b>. '
        f'Sube el archivo de equivalencias para continuar.</div>',
        unsafe_allow_html=True
    )
    archivo_equiv_up = st.file_uploader("Sube equivalencias_libro_rex.xlsx", type=["xlsx"], key="up_equiv")
    if archivo_equiv_up:
        refs_equiv = pd.read_excel(archivo_equiv_up)
        refs_equiv.columns = [str(c).strip() for c in refs_equiv.columns]
        st.markdown('<div class="alert-success">✅ Equivalencias cargadas desde archivo.</div>', unsafe_allow_html=True)
else:
    n_equiv = len([r for _, r in refs_equiv.iterrows()
                   if str(r.get(next((c for c in refs_equiv.columns if "concepto rex" in c.lower()), ""), "") or "").strip() not in ("", "nan")])
    st.markdown(f'<div class="alert-success">✅ Equivalencias cargadas del servidor ({n_equiv} conceptos activos).</div>', unsafe_allow_html=True)

# ── UPLOADERS ──
col_a, col_b = st.columns(2)

with col_a:
    st.markdown("#### 📊 Libro de Remuneraciones (Excel)")
    archivo_libro = st.file_uploader(
        "Sube el Libro xlsx del período",
        type=["xlsx"], key="up_libro",
        help="Libro de Remuneraciones de DH E-Commerce o DH Stores."
    )
    if archivo_libro:
        mes_inferido = inferir_mes_desde_nombre(archivo_libro.name)
        st.markdown(f'<div class="alert-success">✅ Libro cargado: <b>{archivo_libro.name}</b></div>', unsafe_allow_html=True)

    st.markdown("#### 📋 PDF de Liquidaciones (mes a procesar)")
    archivo_pdf = st.file_uploader(
        "Sube el PDF de liquidaciones TALANA del mes a procesar",
        type=["pdf"], key="up_pdf",
        help="PDF con las liquidaciones de sueldo del período actual. Una página = un empleado."
    )
    if archivo_pdf:
        st.markdown(f'<div class="alert-success">✅ PDF cargado: <b>{archivo_pdf.name}</b></div>', unsafe_allow_html=True)

    st.markdown("#### 📚 PDFs históricos (meses anteriores)")
    archivos_pdf_hist = st.file_uploader(
        "Sube los PDFs de todos los meses anteriores",
        type=["pdf"], accept_multiple_files=True, key="up_pdfs_hist",
        help="Usados para calcular renta imponible real de empleados con licencia."
    )
    if archivos_pdf_hist:
        st.markdown(f'<div class="alert-success">✅ {len(archivos_pdf_hist)} PDF(s) histórico(s) cargados.</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-info">ℹ️ Sin PDFs históricos: empleados con licencia usarán el tope AFP como aproximación.</div>', unsafe_allow_html=True)

with col_b:
    st.markdown("#### 👥 Listado de Empleados")
    archivo_empleados = st.file_uploader(
        "Sube listado_empleados.xlsx",
        type=["xlsx"], key="up_empleados",
        help="Contiene: Rut, Empresa, Contrato, AFP, Isapre, Horas, Jornada Parcial."
    )
    if archivo_empleados:
        st.markdown(f'<div class="alert-success">✅ Empleados cargados: <b>{archivo_empleados.name}</b></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-warning">⚠️ Requerido para número de contrato y empresa.</div>', unsafe_allow_html=True)

    st.markdown("#### 🏢 Listado de Empresas")
    archivo_empresas = st.file_uploader(
        "Sube listado_empresas.xlsx",
        type=["xlsx"], key="up_empresas",
        help="Contiene el código Empresa de Rex+ para cada empresa."
    )
    if archivo_empresas:
        st.markdown(f'<div class="alert-success">✅ Empresas cargadas: <b>{archivo_empresas.name}</b></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-warning">⚠️ Requerido para resolver código empresa Rex+.</div>', unsafe_allow_html=True)

    st.markdown("#### 📅 Parámetros mensuales")
    archivo_params_up = st.file_uploader(
        "Sube parametrosMesuales.xlsx (opcional)",
        type=["xlsx"], key="up_params",
        help="Si no subes archivo, se usa la versión guardada en el servidor."
    )
    if archivo_params_up:
        st.markdown(f'<div class="alert-success">✅ Parámetros cargados desde archivo.</div>', unsafe_allow_html=True)
    elif not refs_params.empty:
        st.markdown('<div class="alert-success">✅ Usando parámetros del servidor.</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-error">❌ No hay parámetros disponibles.</div>', unsafe_allow_html=True)

# ── PERÍODO ──
st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)
col_mes, col_fases, _ = st.columns([2, 2, 4])
with col_mes:
    mes_default = mes_inferido if archivo_libro and (mes_inferido := inferir_mes_desde_nombre(archivo_libro.name)) else ""
    mes_proceso = st.text_input(
        "📅 Mes de proceso (aaaa-mm)",
        value=mes_default,
        placeholder="Ej: 2026-01",
        help="Formato año-mes. Se infiere del nombre del archivo si está disponible."
    )
with col_fases:
    usa_fases_ui = st.checkbox(
        "¿Usar Fases Rex+?",
        value=USA_FASES,
        help="Activa si tu configuración Rex+ utiliza fases. Agrega columna Fase=1."
    )

# ── BOTÓN PROCESAR ──
st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)

_listo = bool(archivo_libro and mes_proceso and not refs_equiv.empty)
if not _listo:
    _faltante = []
    if not archivo_libro:       _faltante.append("Libro de Remuneraciones")
    if not mes_proceso:         _faltante.append("Mes de proceso")
    if refs_equiv.empty:        _faltante.append("Equivalencias")
    st.caption(f"⚠️ Falta: {', '.join(_faltante)}")

if st.button("▶ Generar archivo de importación Rex+", disabled=not _listo):
    if not archivo_libro:
        st.markdown('<div class="alert-error">❌ Debes subir el Libro de Remuneraciones.</div>', unsafe_allow_html=True)
        st.stop()
    if not mes_proceso:
        st.markdown('<div class="alert-error">❌ Debes indicar el mes de proceso (aaaa-mm).</div>', unsafe_allow_html=True)
        st.stop()
    if refs_equiv.empty:
        st.markdown('<div class="alert-error">❌ No hay equivalencias cargadas.</div>', unsafe_allow_html=True)
        st.stop()

    with st.spinner("Procesando..."):

        # Cargar Libro
        df_libro, rut_empresa_libro = cargar_libro(archivo_libro.read())
        col_rut_lb = obtener_rut_libro(df_libro)
        if not col_rut_lb:
            st.markdown('<div class="alert-error">❌ No se encontró columna RUT en el Libro.</div>', unsafe_allow_html=True)
            st.stop()
        n_emp_libro = df_libro[col_rut_lb].notna().sum()
        st.markdown(f'<div class="alert-success">✅ Libro leído: <b>{n_emp_libro}</b> empleados, <b>{len(df_libro.columns)}</b> columnas.</div>', unsafe_allow_html=True)
        with st.expander("🔍 Diagnóstico: columnas detectadas en el Libro"):
            st.write("**Columna RUT detectada:**", col_rut_lb)
            st.write("**Primeras 5 filas del Libro:**")
            st.dataframe(df_libro.head(5))
            st.write("**Columnas:**", list(df_libro.columns))

        # Parsear PDF
        pdf_dict = {}
        if archivo_pdf:
            with st.spinner("Parseando PDF de liquidaciones..."):
                pdf_dict, pdf_errors = parsear_pdf_bytes(archivo_pdf.read())
                if pdf_dict is None:
                    st.markdown(f'<div class="alert-error">❌ Error al leer PDF: {pdf_errors}</div>', unsafe_allow_html=True)
                    st.stop()
                if pdf_errors:
                    st.markdown(f'<div class="alert-warning">⚠️ {len(pdf_errors)} página(s) con error en PDF.</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="alert-success">✅ PDF parseado: <b>{len(pdf_dict)}</b> empleados encontrados.</div>', unsafe_allow_html=True)
                # Validar período del PDF vs mes_proceso
                if pdf_dict:
                    primer_emp = next(iter(pdf_dict.values()))
                    periodo_raw = primer_emp.get("periodo_pdf", "")
                    periodo_std = periodo_pdf_a_aaaamm(periodo_raw)
                    if periodo_std and periodo_std != mes_proceso.strip():
                        st.markdown(
                            f'<div class="alert-error">❌ <b>Período del PDF ({periodo_raw} → {periodo_std})</b> no coincide con el mes de proceso ingresado (<b>{mes_proceso}</b>). Verifica que estás subiendo las liquidaciones del mes correcto.</div>',
                            unsafe_allow_html=True
                        )
                        st.stop()
                    elif periodo_std:
                        st.markdown(f'<div class="alert-success">✅ Período del PDF: <b>{periodo_raw}</b> ✔</div>', unsafe_allow_html=True)

            # Parsear PDFs históricos
        pdf_dict_historico = None
        if archivos_pdf_hist:
            with st.spinner(f"Parseando {len(archivos_pdf_hist)} PDF(s) histórico(s)..."):
                pdf_dict_historico = parsear_multiples_pdfs_historicos(archivos_pdf_hist)
                st.markdown(f'<div class="alert-success">✅ Historial: <b>{len(pdf_dict_historico)}</b> empleados con datos históricos.</div>', unsafe_allow_html=True)


        else:
            st.markdown('<div class="alert-warning">⚠️ Sin PDF: Afecto, Id institución, días trab/lic y jornada no estarán disponibles.</div>', unsafe_allow_html=True)

        # Cruzar RUTs entre Libro y PDF (fuera del bloque histórico)
        if pdf_dict:
            ruts_libro = set(df_libro[col_rut_lb].dropna().apply(lambda x: normalizar_rut(str(x))))
            ruts_pdf   = set(pdf_dict.keys())
            sin_pdf    = ruts_libro - ruts_pdf
            if sin_pdf:
                ruts_lista = ", ".join(sorted(sin_pdf))
                st.markdown(
                    f'<div class="alert-warning">⚠️ <b>{len(sin_pdf)}</b> empleado(s) del Libro sin liquidación PDF '
                    f'(días trab/lic y AFP quedarán en 0):<br><small>{ruts_lista}</small></div>',
                    unsafe_allow_html=True
                )

        # Cargar maestros
        df_empleados = None
        df_empresas  = None
        df_params    = refs_params.copy()

        if archivo_empleados:
            if _MODULO_DT_OK:
                df_empleados = _cargar_empleados(archivo_empleados)
            else:
                df_empleados = pd.read_excel(archivo_empleados, header=1)
                df_empleados.columns = [str(c).strip() for c in df_empleados.columns]

        if archivo_empresas:
            # Auto-detect: si la primera fila parece un título (no contiene "empresa"),
            # usar header=1; en caso contrario header=0
            _raw_emp = pd.read_excel(archivo_empresas, header=None, nrows=2, dtype=str)
            _row0 = [str(v).strip().lower() for v in _raw_emp.iloc[0] if str(v).strip().lower() != "nan"]
            _hdr = 1 if not any("empresa" in v for v in _row0) else 0
            df_empresas = pd.read_excel(archivo_empresas, header=_hdr)
            df_empresas.columns = [str(c).strip() for c in df_empresas.columns]

        if archivo_params_up:
            try:
                df_params = pd.read_excel(archivo_params_up, dtype={"mes_Proc": str})
                df_params["mes_Proc"] = df_params["mes_Proc"].astype(str).str.strip()
            except Exception as e:
                st.markdown(f'<div class="alert-warning">⚠️ Error al leer parámetros subidos: {e}. Se usará el del servidor.</div>', unsafe_allow_html=True)

        # Construir equivalencias
        equiv_list = build_equiv_list(refs_equiv)
        if not equiv_list:
            st.markdown('<div class="alert-error">❌ No se pudieron leer las equivalencias. Verifica el formato del archivo.</div>', unsafe_allow_html=True)
            st.stop()

        # Generar archivo de salida
        df_salida, warnings_etl = generar_archivo_salida(
            df_libro, pdf_dict, df_empleados, df_empresas, df_params,
            equiv_list, mes_proceso, archivo_libro.name, usa_fases=usa_fases_ui,
            pdf_dict_historico=pdf_dict_historico,
            rut_empresa_libro=rut_empresa_libro
        )

        if warnings_etl:
            for w in warnings_etl:
                st.markdown(f'<div class="alert-warning">⚠️ {w}</div>', unsafe_allow_html=True)

        if df_salida is None or df_salida.empty:
            st.markdown('<div class="alert-error">❌ No se generaron filas de salida. Verifica el Libro y las equivalencias.</div>', unsafe_allow_html=True)
            st.stop()

        n_filas = len(df_salida)
        n_empl  = df_salida["Id empleado"].nunique() if "Id empleado" in df_salida.columns else 0
        st.markdown(f'<div class="alert-success">✅ Archivo generado: <b>{n_filas:,}</b> filas para <b>{n_empl}</b> empleados.</div>', unsafe_allow_html=True)

        # Previsualización
        with st.expander("🔍 Vista previa (primeras 30 filas)"):
            st.dataframe(df_salida.head(30), use_container_width=True, hide_index=True)

        # Descargar
        excel_bytes = generar_excel_salida(df_salida)
        nombre_salida = f"Rex+_Importacion_{mes_proceso.replace('-', '_')}.xlsx"
        st.download_button(
            label=f"⬇️ Descargar {nombre_salida}",
            data=excel_bytes,
            file_name=nombre_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # Estadísticas
        st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)
        st.markdown("**Resumen por tipo de concepto:**")
        if "Id del concepto" in df_salida.columns and "Monto del concepto" in df_salida.columns:
            resumen = (
                df_salida.groupby("Id del concepto")["Monto del concepto"]
                .agg(["count", "sum"])
                .rename(columns={"count": "Filas", "sum": "Total $"})
                .sort_values("Total $", ascending=False)
                .head(20)
            )
            st.dataframe(resumen, use_container_width=True)

st.divider()
st.caption("Rex+ Tools · Visma · Uso interno")
