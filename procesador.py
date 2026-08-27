import pandas as pd
import numpy as np
from io import BytesIO
import calendar


def es_bisiesto(anio):
    return calendar.isleap(int(anio))


def dias_mes(mes_proc):
    partes = str(mes_proc).split('-')
    anio, mes = int(partes[0]), int(partes[1])
    if mes in [1, 3, 5, 7, 8, 10, 12]:
        return 31
    elif mes in [4, 6, 9, 11]:
        return 30
    elif mes == 2:
        return 29 if es_bisiesto(anio) else 28
    return 30


def safe_float(val):
    try:
        if pd.isna(val):
            return 0.0
        return float(val)
    except:
        return 0.0


def normalizar_rut(rut):
    """Normaliza RUT a formato sin puntos, con guión. Ej: '17.531.760-0' → '17531760-0'"""
    return str(rut).strip().replace('.', '').upper()


def validar_dv_rut(rut_norm):
    """Valida dígito verificador de un RUT normalizado (sin puntos, con guión)."""
    try:
        if '-' not in rut_norm:
            return False
        num_str, dv = rut_norm.rsplit('-', 1)
        num = int(num_str)
        factor = 2
        suma = 0
        while num > 0:
            suma += (num % 10) * factor
            num //= 10
            factor = factor % 7 + 1 if factor < 7 else 2
        resultado = 11 - (suma % 11)
        if resultado == 11:
            dv_calc = '0'
        elif resultado == 10:
            dv_calc = 'K'
        else:
            dv_calc = str(resultado)
        return dv.upper() == dv_calc
    except Exception:
        return False


def procesar_liquidaciones(file_entrada, file_empleados, file_empresas, file_conceptos=None):
    # ── Leer archivos de referencia ───────────────────────────────────────────
    import streamlit as st
    import requests as req

    SUPABASE_URL = st.secrets['SUPABASE_URL']
    SUPABASE_KEY = st.secrets['SUPABASE_KEY']
    hdrs = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}

    def sb_df(table):
        r = req.get(f'{SUPABASE_URL}/rest/v1/{table}?select=*', headers=hdrs)
        return pd.DataFrame(r.json())

    afp_df      = sb_df('inst_afp').set_index('nombre_afp')
    salud_df    = sb_df('inst_salud').set_index('nombre_inst')
    mutuales_df = sb_df('inst_mutuales').set_index('nombre_institucion')
    cajas_df    = sb_df('inst_cajas').set_index('nombre_institucion')
    params_df   = sb_df('parametros_mensuales').set_index('mes_proc')
    params_df.index = params_df.index.astype(str)

    empresas_df  = pd.read_excel(file_empresas, header=1).set_index('Nombre')
    entrada_df   = pd.read_excel(file_entrada)
    empleados_df  = pd.read_excel(file_empleados, header=1)

    # ── Normalizar RUTs del listado de empleados y crear índices ─────────────
    empleados_df['_rut_norm'] = empleados_df['Rut'].apply(normalizar_rut)
    emp_idx         = empleados_df.set_index('_rut_norm')
    ruts_empleados  = set(empleados_df['_rut_norm'])
    sueldo_base_map = emp_idx['Sueldo Base'].apply(safe_float).to_dict()
    tipo_contr_map  = emp_idx['Tipo contr.'].astype(str).to_dict()
    fecha_ces_map   = emp_idx['Fecha inc. Seguro Cesa.'].to_dict()

    # ── Validar que todos los RUTs del archivo de entrada estén en empleados ──
    registros_sin_empleado = []
    registros_rut_invalido = []
    for _, row in entrada_df.iterrows():
        rut_entrada = normalizar_rut(row.get('rut_trabajador', ''))
        if not validar_dv_rut(rut_entrada):
            registros_rut_invalido.append({
                'Rut':                rut_entrada,
                'Número de contrato': row.get('num_contrato', ''),
                'Motivo':             'Dígito verificador inválido',
            })
        elif rut_entrada not in ruts_empleados:
            registros_sin_empleado.append({
                'Rut': rut_entrada,
                'Número de contrato': row.get('num_contrato', '')
            })

    # ── Generar log si hay RUTs con DV inválido ──────────────────────────────
    log_rut_invalido_bytes = None
    if registros_rut_invalido:
        df_rut_inv = pd.DataFrame(registros_rut_invalido)
        buf_rut_inv = BytesIO()
        with pd.ExcelWriter(buf_rut_inv, engine='openpyxl') as writer:
            df_rut_inv.to_excel(writer, index=False, sheet_name='RUTs inválidos')
            ws_ri = writer.sheets['RUTs inválidos']
            from openpyxl.styles import Font, PatternFill, Alignment
            for cell in ws_ri[1]:
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                cell.fill      = PatternFill('solid', start_color='8B0000')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_ri.row_dimensions[1].height = 25
            ws_ri.freeze_panes = 'A2'
            for col in ws_ri.columns:
                ws_ri.column_dimensions[col[0].column_letter].width = 22
        buf_rut_inv.seek(0)
        log_rut_invalido_bytes = buf_rut_inv.read()

    # ── Generar log si hay RUTs no encontrados ────────────────────────────────
    log_bytes = None
    if registros_sin_empleado:
        df_log = pd.DataFrame(registros_sin_empleado)
        buf_log = BytesIO()
        with pd.ExcelWriter(buf_log, engine='openpyxl') as writer:
            df_log.to_excel(writer, index=False, sheet_name='RUTs no encontrados')
            ws = writer.sheets['RUTs no encontrados']
            from openpyxl.styles import Font, PatternFill, Alignment
            hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
            hfill  = PatternFill('solid', start_color='C0392B')
            halign = Alignment(horizontal='center', vertical='center')
            for cell in ws[1]:
                cell.font      = hfont
                cell.fill      = hfill
                cell.alignment = halign
            ws.row_dimensions[1].height = 25
            ws.freeze_panes = 'A2'
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 22
        buf_log.seek(0)
        log_bytes = buf_log.read()

    # ── Mapa Nombre → Concepto(ID) y Nombre → Tipo desde lista de conceptos ──
    nombre_a_id   = {}
    nombre_a_tipo = {}
    hab_exento_nombres   = set()
    hab_afecto_nombres   = set()
    desc_legal_nombres   = set()
    desc_nombres         = set()

    if file_conceptos is not None:
        conc_df = pd.read_excel(file_conceptos, header=1)
        for _, r in conc_df.iterrows():
            nombre_a_id[str(r['Nombre'])]   = str(r['Concepto'])
            nombre_a_tipo[str(r['Nombre'])] = str(r['Tipo'])
            if r['Tipo'] == 'Haber exento':
                hab_exento_nombres.add(str(r['Nombre']))
            if r['Tipo'] == 'Haber afecto':
                hab_afecto_nombres.add(str(r['Nombre']))
            if r['Tipo'] == 'Descuento legal':
                desc_legal_nombres.add(str(r['Nombre']))
            if r['Tipo'] == 'Descuento':
                desc_nombres.add(str(r['Nombre']))

    # ── Columnas fijas del archivo de entrada ─────────────────────────────────
    COLS_FIJAS = {'mes_Proceso', 'rut_trabajador', 'num_contrato', 'nombre_emp',
                  'id_empresa', 'id_afp', 'id_salud', 'id_mutual', 'id_ccaf'}
    concepto_cols = [c for c in entrada_df.columns if c not in COLS_FIJAS]

    # ── Mapeos por Nombre de columna ──────────────────────────────────────────
    CONCEPTOS_AFP_INST = {
        'Cotizacion AFP', 'AFP Reliq meses anteriores', 'Ahorro AFP',
        'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
        'Trabajo Pesado Empleado', 'APVI Ahorro voluntario mensual',
        'APVC Ahorro voluntario colectivo', 'APVI Deposito Convenido',
        'Afiliado Voluntario Cotizacion', 'Afiliado Voluntario Ahorro',
        'Trabajo pesado Empl Reliq anteriores', 'Seguro Invalidez y Sobrevivencia',
        'SIS Reliq meses anteriores', 'CESANTIA CI Reliq meses anteriores',
        'CESANTIA SOL Reliq meses anteriores', 'Seguro de Cesantia CI',
        'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
        'Aporte FAPP Compensación Expectativa de Vida',
        'Seguro de Cesantia Solidario', 'Trabajo Pesado',
        'Trabajo pesado Reliq anteriores', 'APVC - Aporte Empleador'
    }
    CONCEPTOS_SALUD_INST  = {'Cotizacion SALUD', 'SALUD Reliq meses anteriores', 'Aporte Seguro Salud'}
    CONCEPTOS_MUTUAL_INST = {'Mutual', 'MUTUAL Reliq meses anteriores'}
    CONCEPTOS_CCAF_INST   = {'Aporte a CCAF', 'CCAF Reliq meses anteriores'}

    CONCEPTOS_AFP_AFECTO = {
        'Cotizacion AFP', 'AFP Reliq meses anteriores', 'Ahorro AFP',
        'Trabajo Pesado', 'Trabajo pesado Reliq anteriores',
        'Trabajo Pesado Empleado', 'Trabajo pesado Empl Reliq anteriores',
        'Mutual', 'MUTUAL Reliq meses anteriores',
        'Seguro Invalidez y Sobrevivencia', 'SIS Reliq meses anteriores',
        'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
        'Aporte FAPP Compensación Expectativa de Vida', 'APVC - Aporte Empleador'
    }
    CONCEPTOS_CES_AFECTO = {
        'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
        'Seguro de Cesantia CI', 'CESANTIA CI Reliq meses anteriores',
        'Seguro de Cesantia Solidario', 'CESANTIA SOL Reliq meses anteriores',
    }

    cotizacion_afp_idx = concepto_cols.index('Cotizacion AFP') if 'Cotizacion AFP' in concepto_cols else len(concepto_cols)
    haberes_cols = concepto_cols[:cotizacion_afp_idx]

    filas_salida        = []
    registros_descuadre = []
    registros_ces       = []
    registros_ult_imp   = []

    for _, row in entrada_df.iterrows():
        mes_proc      = str(row.get('mes_Proceso', ''))
        rut           = row.get('rut_trabajador', '')
        num_cont      = row.get('num_contrato', '')
        nombre_emp    = row.get('nombre_emp', '')
        id_empresa    = row.get('id_empresa', '')
        id_afp_val    = str(row.get('id_afp', ''))
        id_salud_val  = str(row.get('id_salud', ''))
        id_mutual_val = str(row.get('id_mutual', ''))
        id_ccaf_val   = str(row.get('id_ccaf', ''))

        rut_norm      = normalizar_rut(rut)

        params_row = params_df.loc[mes_proc] if mes_proc in params_df.index else None
        tope_afp   = safe_float(params_row['tope_imp_pesos_afp']) if params_row is not None else 0
        tope_ces   = safe_float(params_row['tope_ces_pesos'])     if params_row is not None else 0
        tope_salud = safe_float(params_row['tope_salud_pesos'])   if params_row is not None else 0

        def lookup_param(col):
            if params_row is not None and col in params_df.columns:
                return safe_float(params_row[col])
            return 0

        def lookup_afp_id():
            return afp_df.loc[id_afp_val, 'id_afp'] if id_afp_val in afp_df.index else ''

        def lookup_salud_id():
            return salud_df.loc[id_salud_val, 'id_inst'] if id_salud_val in salud_df.index else ''

        def lookup_mutual_id():
            return mutuales_df.loc[id_mutual_val, 'id_institucion'] if id_mutual_val in mutuales_df.index else ''

        def lookup_ccaf_id():
            return cajas_df.loc[id_ccaf_val, 'id_institucion'] if id_ccaf_val in cajas_df.index else ''

        total_hab_afecto = sum(safe_float(row.get(c, 0)) for c in haberes_cols if c in hab_afecto_nombres or c not in hab_exento_nombres)
        total_hab_exento = sum(safe_float(row.get(c, 0)) for c in haberes_cols if c in hab_exento_nombres)
        imponible        = min(total_hab_afecto, tope_afp)

        val_afp            = safe_float(row.get('Cotizacion AFP', 0))
        val_ces_empleado   = safe_float(row.get('Seguro de Cesantia', 0))
        val_trab_pesa_empl = safe_float(row.get('Trabajo Pesado Empleado', 0))
        val_isapre         = safe_float(row.get('Cotizacion SALUD', 0))
        val_sueldo_base    = safe_float(row.get('sueldoBase', 0))
        val_licencia       = safe_float(row.get('licenciaDias', 0)) if 'licenciaDias' in row.index else 0

        dias_lic        = val_licencia if val_licencia > 0 else 0
        total_dias      = dias_mes(mes_proc)
        dias_perm_falt  = safe_float(row.get('Dias Perm y Faltas', 0))
        dias_lic_med    = safe_float(row.get('Dias Lic. Med.', 0))
        ausencias       = dias_perm_falt + dias_lic_med

        anio_proc, mes_proc_num = int(mes_proc[:4]), int(mes_proc[5:7])
        if mes_proc_num == 2:
            dias_trab = total_dias - ausencias
        elif dias_lic_med > 0:
            dias_trab = 31 - ausencias
        else:
            dias_trab = 30 - ausencias

        dias_trab  = max(0, round(dias_trab, 2))
        monto_init = round((val_sueldo_base / dias_trab) * 30, 2) if dias_trab > 0 else 0

        # ── Validación Ult. Imp 30 dias ───────────────────────────────────────
        ult_imp_30 = safe_float(row.get('Ult. Imp 30 dias', 0))
        if dias_lic_med > 0 and ult_imp_30 == 0:
            registros_ult_imp.append({
                'Rut':                rut,
                'Número de contrato': num_cont,
                'Dias Lic. Med.':     dias_lic_med,
                'Ult. Imp 30 dias':   ult_imp_30,
                'Observación':        '"Ult. Imp 30 dias" debe ser > 0 cuando "Dias Lic. Med." > 0',
            })

        # ── Aportes empleador — calcular si vienen en blanco o cero ──────────
        tipo_contr    = tipo_contr_map.get(rut_norm, '')
        fecha_ces_raw = fecha_ces_map.get(rut_norm, None)

        # Trabajo Pesado
        val_trab_pesa = safe_float(row.get('Trabajo Pesado', 0))
        if val_trab_pesa == 0:
            val_trab_pesa = val_trab_pesa_empl

        # Aporte a CCAF — solo cuando Cotizacion SALUD > 0
        val_ccaf = safe_float(row.get('Aporte a CCAF', 0))
        if val_ccaf == 0 and val_isapre > 0:
            val_ccaf = round((lookup_param('aporte_ccaf') / 100) * imponible, 2)

        # Seguro Invalidez y Sobrevivencia (SIS)
        val_sis = safe_float(row.get('Seguro Invalidez y Sobrevivencia', 0))
        if val_sis == 0:
            val_sis = round((lookup_param('sis') / 100) * imponible, 2)

        # ── Base con licencia médica ──────────────────────────────────────────
        base_lic = imponible + (ult_imp_30 / 30 * dias_lic_med) if dias_lic_med > 0 else imponible

        # ── CesAporteCi y cesAporteSol ────────────────────────────────────────
        val_ces_ci  = safe_float(row.get('CesAporteCi', 0))
        val_ces_sol = safe_float(row.get('cesAporteSol', 0))
        if val_ces_ci == 0 or val_ces_sol == 0:
            if tipo_contr == 'I':
                val_ces_ci  = round(0.016 * base_lic, 2)
                val_ces_sol = round(0.008 * base_lic, 2)
            elif tipo_contr in ('F', 'O'):
                val_ces_ci  = round(0.028 * base_lic, 2)
                val_ces_sol = round(0.002 * base_lic, 2)

        # ── aporteFAPPCEV ─────────────────────────────────────────────────────
        val_fapp = safe_float(row.get('aporteFAPPCEV', 0))
        if val_fapp == 0:
            val_fapp = round(0.009 * base_lic, 2)

        # ── Mutual con licencia ───────────────────────────────────────────────
        val_mutual = safe_float(row.get('Mutual', 0))
        if val_mutual == 0:
            try:
                cot_mutual = safe_float(empresas_df.loc[str(id_empresa), 'Cotización Mutual']) if str(id_empresa) in empresas_df.index else 0
                if dias_lic_med > 0:
                    val_mutual = round((cot_mutual / 100 * imponible) + (ult_imp_30 / 30 * dias_lic_med), 2)
                else:
                    val_mutual = round(cot_mutual / 100 * imponible, 2)
            except Exception:
                val_mutual = 0

        # Actualizar valores en row
        row = row.copy()
        row['Trabajo Pesado']                    = val_trab_pesa
        row['Aporte a CCAF']                     = val_ccaf
        row['Mutual']                            = val_mutual
        row['Seguro Invalidez y Sobrevivencia']  = val_sis
        row['CesAporteCi']                       = val_ces_ci
        row['cesAporteSol']                      = val_ces_sol
        row['aporteFAPPCEV']                     = val_fapp

        total_hab      = sum(safe_float(row.get(c, 0)) for c in concepto_cols if c in hab_afecto_nombres or c in hab_exento_nombres)
        total_desc     = sum(safe_float(row.get(c, 0)) for c in concepto_cols if c in desc_legal_nombres or c in desc_nombres)
        liquido_calc   = total_hab - total_desc
        liquido_ingres = safe_float(row.get('totalesEmpl', 0))
        diferencia     = liquido_ingres - liquido_calc

        if diferencia != 0:
            registros_descuadre.append({
                'Rut':               rut,
                'Número de contrato': num_cont,
                'Líquido calculado': liquido_calc,
                'Líquido ingresado': liquido_ingres,
                'Diferencia':        diferencia,
            })
            continue

        # ── Validación seguro de cesantía ─────────────────────────────────────
        val_ces = safe_float(row.get('Seguro de Cesantia', 0))

        # Validación 1: Tipo contrato I debe tener valor en Seguro de Cesantia
        if tipo_contr == 'I' and val_ces == 0:
            registros_ces.append({
                'Rut':                rut,
                'Número de contrato': num_cont,
                'Validación fallida': 'Contrato tipo I sin valor en Seguro de Cesantia',
            })

        # Validación 2: Más de 11 años desde Fecha inc. Seguro Cesa. → no debe venir valor
        if fecha_ces_raw is not None and val_ces > 0:
            try:
                if pd.notna(fecha_ces_raw):
                    fecha_ces_dt = pd.to_datetime(fecha_ces_raw)
                    anio_proc, mes_proc_num = int(mes_proc[:4]), int(mes_proc[5:7])
                    fecha_proc_dt = pd.Timestamp(year=anio_proc, month=mes_proc_num, day=1)
                    años_diff = (fecha_proc_dt - fecha_ces_dt).days / 365.25
                    if años_diff > 11:
                        registros_ces.append({
                            'Rut':                rut,
                            'Número de contrato': num_cont,
                            'Validación fallida': f'Más de 11 años en Seguro Cesantía ({años_diff:.1f} años) — no debería venir valor',
                        })
            except Exception:
                pass

        for concepto_nombre in concepto_cols:
            monto = safe_float(row.get(concepto_nombre, 0))
            if monto <= 0:
                continue

            id_concepto    = nombre_a_id.get(concepto_nombre, concepto_nombre)
            monto_concepto = monto

            if concepto_nombre in {'totalesEmpl', 'Sueldo liquido'}:
                afecto = total_hab_afecto + total_hab_exento
            elif concepto_nombre in CONCEPTOS_AFP_AFECTO:
                afecto = imponible
            elif concepto_nombre in CONCEPTOS_CES_AFECTO:
                afecto = min(imponible, tope_ces)
            elif concepto_nombre in {'Impuesto mensual', 'IMPUESTO Reliq meses anteriores'}:
                afecto = imponible - (val_afp + val_ces_empleado + val_trab_pesa_empl + min(val_isapre, tope_salud))
            elif concepto_nombre in {'SALUD Reliq meses anteriores'}:
                afecto = ''
            else:
                afecto = 0

            if concepto_nombre in CONCEPTOS_AFP_INST:
                id_inst = lookup_afp_id()
            elif concepto_nombre in CONCEPTOS_SALUD_INST:
                id_inst = lookup_salud_id()
            elif concepto_nombre in CONCEPTOS_MUTUAL_INST:
                id_inst = lookup_mutual_id()
            elif concepto_nombre in CONCEPTOS_CCAF_INST:
                id_inst = lookup_ccaf_id()
            else:
                id_inst = ''

            if concepto_nombre == 'Cotizacion AFP':
                cot_jub = safe_float(afp_df.loc[id_afp_val, 'cot_afp']) if id_afp_val in afp_df.index else ''
            elif concepto_nombre == 'Cotizacion SALUD':
                cot_jub = monto_concepto
            elif concepto_nombre == 'Aporte a CCAF':
                cot_jub = lookup_param('aporte_ccaf')
            elif concepto_nombre == 'Mutual':
                cot_jub = safe_float(empresas_df.loc[str(nombre_emp), 'Cotización Mutual']) if str(nombre_emp) in empresas_df.index else ''
            elif concepto_nombre == 'Seguro Invalidez y Sobrevivencia':
                cot_jub = lookup_param('sis')
            elif concepto_nombre == 'Aporte AFP Empleador':
                cot_jub = lookup_param('aporte_afp')
            elif concepto_nombre == 'Aporte FAPP Compensación Expectativa de Vida':
                cot_jub = lookup_param('seg_social_exp_vida')
            else:
                cot_jub = ''

            dias_licencias  = monto if concepto_nombre == 'licenciaDias' and monto > 0 else 0
            dias_trabajados = dias_trab

            if concepto_nombre == 'Impuesto mensual':
                total_rebajas = val_afp + val_ces_empleado + val_trab_pesa_empl + min(val_isapre, tope_salud)
            else:
                total_rebajas = 0

            if concepto_nombre == 'Impuesto mensual':
                rentas_no_grav = total_hab_exento
            else:
                rentas_no_grav = 0

            filas_salida.append({
                'Fecha de proceso':          mes_proc,
                'Id empleado':               rut,
                'Número de contrato':        num_cont,
                'Id del concepto':           id_concepto,
                'Monto del concepto':        monto_concepto,
                'Afecto':                    afecto,
                'Id de institución':         id_inst,
                'Cotización de jubilación':  cot_jub,
                'Días de licencias':         dias_licencias,
                'Días trabajados':           dias_trabajados,
                'Fecha de aplicación':       'x',
                'Empresa':                   id_empresa,
                'Total de rebajas por LLSS': total_rebajas,
                'Rentas no gravadas':        rentas_no_grav,
                'Rebaja por zona extrema':   0,
                'Jornada':                   'C',
                'Días de vacaciones':        '',
                'Monto Init':                monto_init,
                'Fase':                      1,
            })

    df_salida = pd.DataFrame(filas_salida)
    n_trabajadores = len(entrada_df)
    n_filas        = len(df_salida)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_salida.to_excel(writer, index=False, sheet_name='Liquidaciones detalladas')
        ws = writer.sheets['Liquidaciones detalladas']
        from openpyxl.styles import Font, PatternFill, Alignment
        hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        hfill  = PatternFill('solid', start_color='1a2744')
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[1]:
            cell.font      = hfont
            cell.fill      = hfill
            cell.alignment = halign
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    buf.seek(0)

    # ── Log de descuadre de líquido ───────────────────────────────────────────
    log_descuadre_bytes = None
    if registros_descuadre:
        df_desc = pd.DataFrame(registros_descuadre)
        buf_desc = BytesIO()
        with pd.ExcelWriter(buf_desc, engine='openpyxl') as writer:
            df_desc.to_excel(writer, index=False, sheet_name='Descuadre líquido')
            ws2 = writer.sheets['Descuadre líquido']
            hfill2 = PatternFill('solid', start_color='C0392B')
            for cell in ws2[1]:
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                cell.fill      = hfill2
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[1].height = 25
            ws2.freeze_panes = 'A2'
            for col in ws2.columns:
                ws2.column_dimensions[col[0].column_letter].width = 22
        buf_desc.seek(0)
        log_descuadre_bytes = buf_desc.read()

    # ── Log de validación seguro de cesantía ──────────────────────────────────
    log_ces_bytes = None
    if registros_ces:
        df_ces = pd.DataFrame(registros_ces)
        buf_ces = BytesIO()
        with pd.ExcelWriter(buf_ces, engine='openpyxl') as writer:
            df_ces.to_excel(writer, index=False, sheet_name='Validación cesantía')
            ws3 = writer.sheets['Validación cesantía']
            hfill3 = PatternFill('solid', start_color='E67E22')
            for cell in ws3[1]:
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                cell.fill      = hfill3
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws3.row_dimensions[1].height = 25
            ws3.freeze_panes = 'A2'
            for col in ws3.columns:
                ws3.column_dimensions[col[0].column_letter].width = 28
        buf_ces.seek(0)
        log_ces_bytes = buf_ces.read()

    # ── Log de validación Ult. Imp 30 dias ───────────────────────────────────
    log_ult_imp_bytes = None
    if registros_ult_imp:
        df_ult = pd.DataFrame(registros_ult_imp)
        buf_ult = BytesIO()
        with pd.ExcelWriter(buf_ult, engine='openpyxl') as writer:
            df_ult.to_excel(writer, index=False, sheet_name='Ult. Imp 30 dias')
            ws_ult = writer.sheets['Ult. Imp 30 dias']
            for cell in ws_ult[1]:
                cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
                cell.fill      = PatternFill('solid', start_color='E67E22')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_ult.row_dimensions[1].height = 25
            ws_ult.freeze_panes = 'A2'
            for col in ws_ult.columns:
                ws_ult.column_dimensions[col[0].column_letter].width = 25
        buf_ult.seek(0)
        log_ult_imp_bytes = buf_ult.read()

    return buf.read(), n_filas, n_trabajadores, registros_sin_empleado, log_bytes, registros_descuadre, log_descuadre_bytes, registros_ces, log_ces_bytes, registros_rut_invalido, log_rut_invalido_bytes, registros_ult_imp, log_ult_imp_bytes
